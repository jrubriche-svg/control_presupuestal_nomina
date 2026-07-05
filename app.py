import streamlit as st
import pandas as pd
import hashlib
import time
from datetime import datetime
from pathlib import Path

# =============================================================================
# FUNCIONES PARA EXPORTAR A EXCEL CON FORMATOS
# =============================================================================
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter

def aplicar_formato_excel(writer, resumen, nombre_hoja, tipo_tabla="RP"):
    """
    Aplica formatos de estilo a una hoja de Excel
    
    Args:
        writer: ExcelWriter
        resumen: DataFrame a exportar
        nombre_hoja: Nombre de la hoja
        tipo_tabla: "RP" o "SGP" para diferentes estilos
    """
    # Exportar DataFrame primero
    resumen.to_excel(writer, sheet_name=nombre_hoja, index=True)
    
    # Obtener la hoja
    worksheet = writer.sheets[nombre_hoja]
    
    # Definir estilos
    estilo_encabezado = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    estilo_encabezado_fuente = Font(color="FFFFFF", bold=True, size=11)
    estilo_total = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    estilo_total_final = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    estilo_total_general = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    estilo_fila_par = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    estilo_fila_impar = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    estilo_borde = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    estilo_fuente_normal = Font(size=10)
    estilo_fuente_numero = Font(size=10, name='Courier New', bold=True)
    estilo_alineacion_centro = Alignment(horizontal='center', vertical='center', wrap_text=True)
    estilo_alineacion_izquierda = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Determinar filas totales según el tipo de tabla
    if tipo_tabla == "RP":
        filas_totales = {
            "SUELDOS": estilo_total,
            "PARAFISCALES": estilo_total_final,
            "FOMAG": estilo_total_general,
            "TOTAL_DOC_RP": estilo_total,
            "SUELDOS_ORIENTADORES": estilo_total,
            "PARAFISCALES_ORIENTADORES": estilo_total_final,
            "FOMAG_ORIENTADORES": estilo_total_general,
            "TOTAL_DOC_ORIENTADORES": estilo_total,
            "SUELDOS_PBM": estilo_total,
            "PARAFISCALES_PBM": estilo_total_final,
            "FOMAG_PBM": estilo_total_general,
            "TOTAL_DOC_PBM": estilo_total,
            "SUELDOS": estilo_total,
            "PARAFISCALES": estilo_total_final,
            "FOMAG": estilo_total_general,
            "TOTAL_PRIMERA_INFANCIA": estilo_total
        }
    else:  # SGP
        filas_totales = {
            "TOTAL_DOC_SGP": estilo_total,
            "TOTAL_SGP_P8033": estilo_total_final,
            "TOTAL_RP_P8033": estilo_total_final,
            "TOTAL_GENERAL": estilo_total_general
        }
    
    # Aplicar formatos a todas las celdas
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                   min_col=1, max_col=worksheet.max_column):
        for cell in row:
            # Aplicar bordes a todas las celdas
            cell.border = estilo_borde
            
            # Aplicar formato a encabezados (fila 1)
            if cell.row == 1:
                cell.fill = estilo_encabezado
                cell.font = estilo_encabezado_fuente
                cell.alignment = estilo_alineacion_centro
            else:
                # Formato para columnas de texto (primera columna)
                if cell.column == 1:
                    cell.alignment = estilo_alineacion_izquierda
                    cell.font = estilo_fuente_normal
                    
                    # Verificar si es fila total
                    fila_nombre = resumen.index[cell.row-2] if cell.row-2 < len(resumen.index) else ""
                    if fila_nombre in filas_totales:
                        cell.fill = filas_totales[fila_nombre]
                        cell.font = Font(color="FFFFFF", bold=True, size=10)
                    else:
                        # Filas alternadas
                        if cell.row % 2 == 0:
                            cell.fill = estilo_fila_par
                        else:
                            cell.fill = estilo_fila_impar
                
                # Formato para columnas numéricas
                else:
                    cell.alignment = estilo_alineacion_centro
                    cell.font = estilo_fuente_numero
                    
                    # Aplicar formato de moneda a celdas con números
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '"$"#,##0'
                    
                    # Verificar si es fila total
                    fila_nombre = resumen.index[cell.row-2] if cell.row-2 < len(resumen.index) else ""
                    if fila_nombre in filas_totales:
                        cell.fill = filas_totales[fila_nombre]
                        cell.font = Font(color="FFFFFF", bold=True, size=10, name='Courier New')
                    else:
                        # Filas alternadas
                        if cell.row % 2 == 0:
                            cell.fill = estilo_fila_par
                        else:
                            cell.fill = estilo_fila_impar
    
    # Ajustar ancho de columnas
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Ajustar ancho, con límites
        adjusted_width = min(max_length + 2, 50)
        
        # Para columnas de texto largas (como NOMBRE y CONCEPTO)
        if column_letter in ['B', 'C'] and adjusted_width < 30:
            adjusted_width = 30
        
        worksheet.column_dimensions[column_letter].width = adjusted_width

def exportar_a_excel_formateado(lista_resumenes, nombres_hojas=None, tipos_tablas=None):
    """
    Exporta múltiples DataFrames a un archivo Excel con formatos de estilo
    
    Args:
        lista_resumenes: Lista de DataFrames a exportar
        nombres_hojas: Lista de nombres para cada hoja
        tipos_tablas: Lista de tipos ("RP" o "SGP") para cada hoja
    
    Returns:
        BytesIO object con el archivo Excel formateado
    """
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for i, resumen in enumerate(lista_resumenes):
            if resumen is not None and not resumen.empty:
                # Usar nombre de hoja personalizado o por defecto
                nombre_hoja = nombres_hojas[i] if nombres_hojas else f'Hoja_{i+1}'
                nombre_hoja = nombre_hoja[:31]  # Límite de Excel
                
                # Tipo de tabla (por defecto "RP")
                tipo_tabla = tipos_tablas[i] if tipos_tablas and i < len(tipos_tablas) else "RP"
                
                # Aplicar formatos
                aplicar_formato_excel(writer, resumen, nombre_hoja, tipo_tabla)
    
    output.seek(0)
    return output

# =============================================================================
# CONFIGURACIÓN
# =============================================================================
st.set_page_config(page_title="PRESUPUESTOS", page_icon="⭐", layout="wide")

# =============================================================================
# ESTILOS PERSONALIZADOS
# =============================================================================
def cargar_estilos():
    st.markdown("""
        <style>
        body {
            background-color: #f5f5f5;
            font-family: 'Arial', sans-serif;
        }
        .header {
            background-color: #D50000;
            padding: 20px;
            border-radius: 10px;
            color: white;
            text-align: center;
            margin-bottom: 20px;
        }
        .subtitulo {
            color: #333;
            font-size: 18px;
            margin-top: 20px;
            margin-bottom: 10px;
            font-weight: bold;
        }
        .tabla-container {
            border: 3px solid #b30000;
            border-radius: 10px;
            overflow: hidden;
            margin: 20px 0;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        }
        .tabla-personalizada {
            width: 100%;
            border-collapse: collapse;
            font-family: Arial, sans-serif;
            font-size: 13px;
        }
        .tabla-personalizada th {
            background-color: #b30000;
            color: white;
            font-weight: bold;
            padding: 8px 6px;
            text-align: center;
            border: 2px solid #8b0000;
            font-size: 13px;
            line-height: 1.2;
        }
        .tabla-personalizada td {
            padding: 6px 4px;
            text-align: center;
            border: 2px solid #ddd;
            line-height: 1.2;
        }
        .tabla-personalizada tr:nth-child(even):not(.fila-total) {
            background-color: #f9f9f9;
        }
        .tabla-personalizada tr:nth-child(odd):not(.fila-total) {
            background-color: white;
        }
        .tabla-personalizada tr:hover:not(.fila-total) {
            background-color: #f0f0f0;
            transition: background-color 0.3s;
        }
        .fila-total {
            background-color: #ff6b6b !important;
            font-weight: bold;
            color: #000;
            font-size: 13px;
        }
        .fila-total-final {
            background-color: #4CAF50 !important;
            font-weight: bold;
            color: white;
            font-size: 14px;
        }
        .fila-total-general {
            background-color: #2196F3 !important;
            font-weight: bold;
            color: white;
            font-size: 15px;
        }
        .encabezado-fila {
            background-color: #b30000;
            color: white;
            font-weight: bold;
            text-align: left;
            border: 2px solid #8b0000;
            padding: 6px 8px !important;
            line-height: 1.2;
        }
        .numero {
            font-family: 'Courier New', monospace;
            font-weight: bold;
            font-size: 12px;
        }
        .titulo-tabla {
            color: #b30000;
            text-align: center;
            margin: 20px 0 10px 0;
            font-size: 20px;
            font-weight: bold;
        }
        .boton-menu {
            background-color: #D50000;
            color: white;
            padding: 25px;
            border-radius: 15px;
            text-align: center;
            font-size: 22px;
            font-weight: bold;
            cursor: pointer;
            transition: all 0.3s;
            border: none;
            width: 100%;
            margin: 15px 0;
            box-shadow: 0 4px 8px rgba(0, 0, 0, 0.2);
        }
        .boton-menu:hover {
            background-color: #b30000;
            transform: translateY(-2px);
            box-shadow: 0 6px 12px rgba(0, 0, 0, 0.3);
        }
        .contenedor-botones {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 30px;
            margin: 20px 0;
            max-width: 800px;
            margin-left: auto;
            margin-right: auto;
        }
        .pantalla-inicial {
            text-align: center;
            padding: 10px;
        }
        .boton-volver {
            background-color: #666;
            color: white;
            padding: 10px 20px;
            border-radius: 5px;
            border: none;
            cursor: pointer;
            margin-bottom: 20px;
        }
        .boton-volver:hover {
            background-color: #555;
        }
        .contenedor-logos {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 10px;
        }
        .logo {
            max-width: 250px;
            height: auto;
        }
        .boton-actualizar {
            background-color: #4CAF50;
            color: white;
            padding: 10px 20px;
            border-radius: 5px;
            border: none;
            cursor: pointer;
            font-weight: bold;
            margin: 10px 0;
        }
        .boton-actualizar:hover {
            background-color: #45a049;
        }
         /* 🔹 NUEVO ESTILO PARA COLUMNAS DE TEXTO LARGO */
        .texto-reducido {
            font-size: 11px !important;
            line-height: 1.1 !important;
            padding: 4px 2px !important;
            white-space: normal !important;
            word-wrap: break-word !important;
            max-width: 150px !important;
        }
        
        /* Opcional: Estilo para encabezados de estas columnas */
        .encabezado-texto {
            font-size: 11px !important;
            padding: 6px 2px !important;
            white-space: normal !important;
        }    
        </style>
    """, unsafe_allow_html=True)

def mostrar_fecha_actualizacion():
    """Muestra la fecha de actualización de los datos en un formato estilizado"""
    fecha = st.session_state.get('fecha_actualizacion', 'No disponible')
    st.markdown(f"""
        <div style="
            background: linear-gradient(135deg, #e8f5e9 0%, #c8e6c9 100%);
            border-left: 5px solid #2E7D32;
            border-radius: 8px;
            padding: 12px 20px;
            margin: 10px 0 20px 0;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            display: flex;
            align-items: center;
            gap: 12px;
        ">
            <span style="font-size: 22px;">📅</span>
            <div>
                <span style="font-weight: 600; color: #1B5E20; font-size: 15px;">Última actualización:</span>
                <span style="color: #2E7D32; font-size: 15px; margin-left: 6px;">{fecha}</span>
            </div>
        </div>
    """, unsafe_allow_html=True)
# =============================================================================
# CARGAR DATOS DESDE GOOGLE SHEETS - VERSIÓN PRODUCCIÓN
# =============================================================================

@st.cache_data(ttl=30, show_spinner="📥 Sincronizando con Google Sheets...")
def cargar_datos_originales(_fuerza_actualizacion=False):
    """
    Carga datos desde Google Sheets con detección automática de cambios.
    Args:
        _fuerza_actualizacion: Forzar recarga ignorando cache
    Returns:
        DataFrame con los datos o None si hay error
    """
    try:
        # URL DE TU GOOGLE SHEETS
        BASE_URL = "https://docs.google.com/spreadsheets/d/1MK6NNx5YEqo_19xdSwpXg_WRYd52GPTpFPeVMYeZCNo/export?format=csv&gid=0"

        # Estrategia: Si es forzado, agregamos timestamp único
        if _fuerza_actualizacion:
            timestamp = int(time.time() * 1000)
            url = f"{BASE_URL}&_={timestamp}"
            st.toast("🔄 Actualización forzada solicitada", icon="📡")
        else:
            url = BASE_URL
        
        # Leer datos
        df = pd.read_csv(url)
        
        # VALIDACIÓN CRÍTICA
        if df.empty:
            st.warning("⚠️ La hoja está vacía o no es accesible")
            return None
        
        # 🔹 EXTRAER FECHA DE ACTUALIZACIÓN - CORREGIDO
        if "FECHA" in df.columns:
            # 🔹 PASO 1: Forzar el formato correcto de fecha
            # Si la fecha viene como "4/07/2026" (día/mes/año)
            # Usamos dayfirst=True para que pandas lo interprete correctamente
            
            try:
                # Intento 1: Formato día/mes/año (DD/MM/YYYY)
                df["FECHA"] = pd.to_datetime(df["FECHA"], format="%d/%m/%Y", errors="coerce")
            except:
                try:
                    # Intento 2: Formato día/mes/año con día de 2 dígitos (04/07/2026)
                    df["FECHA"] = pd.to_datetime(df["FECHA"], format="%d/%m/%Y", errors="coerce")
                except:
                    try:
                        # Intento 3: Formato día/mes/año con horas (04/07/2026 14:30)
                        df["FECHA"] = pd.to_datetime(df["FECHA"], format="%d/%m/%Y %H:%M", errors="coerce")
                    except:
                        try:
                            # Intento 4: Formato día/mes/año con horas y minutos (04/07/2026 14:30:00)
                            df["FECHA"] = pd.to_datetime(df["FECHA"], format="%d/%m/%Y %H:%M:%S", errors="coerce")
                        except:
                            # Intento 5: Dejar que pandas intente adivinar automáticamente
                            # pero especificando dayfirst=True
                            df["FECHA"] = pd.to_datetime(df["FECHA"], dayfirst=True, errors="coerce")
            
            # 🔹 PASO 2: Obtener la fecha más reciente
            fecha_reciente = df["FECHA"].max()
            
            if pd.notna(fecha_reciente):
                # 🔹 PASO 3: Guardar en session state con formato en letras
                # Opción 1: Solo fecha (ej: "04 de Julio de 2026")
                st.session_state.fecha_actualizacion = fecha_reciente.strftime("%d de %B de %Y")
                
                # Opción 2: Con día de semana (ej: "Lunes 04 de Julio de 2026")
                # st.session_state.fecha_actualizacion = fecha_reciente.strftime("%A %d de %B de %Y")
                
                # Opción 3: Con hora (ej: "04 de Julio de 2026 a las 02:30 PM")
                # st.session_state.fecha_actualizacion = fecha_reciente.strftime("%d de %B de %Y a las %I:%M %p")
                
                # 🔹 Para depuración: Mostrar en consola la fecha correcta
                print(f"✅ Fecha correcta: {st.session_state.fecha_actualizacion}")
            else:
                st.session_state.fecha_actualizacion = "No disponible"
        else:
            st.session_state.fecha_actualizacion = "Columna FECHA no encontrada"
        
        # DETECTOR DE CAMBIOS INTELIGENTE
        contenido_actual = df.to_csv(index=False).encode('utf-8')
        hash_actual = hashlib.md5(contenido_actual).hexdigest()
        
        # Guardar en session state para comparar
        if 'data_hash' not in st.session_state:
            st.session_state.data_hash = hash_actual
        elif st.session_state.data_hash != hash_actual:
            st.session_state.data_hash = hash_actual
            st.toast("✅ Nuevos datos detectados", icon="🆕")
            st.sidebar.success("📊 Datos actualizados")
        
        # PREPROCESAMIENTO ESENCIAL
        if "Codigo" in df.columns:
            # Crear Codigo_O si no existe
            if "Codigo_O" not in df.columns:
                df.insert(0, "Codigo_O", df["Codigo"].where(
                    df["Codigo"].astype(str).str.startswith("O")
                ).ffill())
            
            # Asegurar columna Concepto de gasto
            if "Concepto de gasto" in df.columns:
                df["Concepto de gasto"] = df["Concepto de gasto"].ffill()
        
        # Indicador sutil de éxito
        if not _fuerza_actualizacion:
            hora_actual = datetime.now().strftime("%H:%M:%S")
            st.sidebar.caption(f"Última sync: {hora_actual}")
            # 🔹 Mostrar fecha en letras en la barra lateral
            st.sidebar.caption(f"📅 {st.session_state.get('fecha_actualizacion', 'Desconocida')}")
        
        return df
        
    except Exception as e:
        # ERROR GRACEFUL - No colapsar la app
        st.error(f"❌ Error de conexión: {str(e)[:100]}...")
        
        # CARGAR VERSIÓN EN CACHE COMO RESCATE
        if 'df_backup' in st.session_state:
            st.warning("⚠️ Usando datos en cache (fuente no disponible)")
            return st.session_state.df_backup
        
        return None

# =============================================================================
# FUNCIÓN DE PROCESAMIENTO TABLERO PRINCIPAL
# =============================================================================
def procesar_datos_sgp(fuerza_actualizacion=False):
    """Función específica para procesar datos SGP"""
    # Cargar datos
    df = cargar_datos_originales(_fuerza_actualizacion=fuerza_actualizacion)
    
    if df is None:
        return None, None
    
    try:        
        # 🔹 CALCULAR 'ultimos_cuatro' LOCALMENTE (IMPORTANTE)
        ultimos_cuatro = pd.to_numeric(
            df["Codigo_O"].astype(str).str[-4:], 
            errors="coerce"
        )
        
        # 🔹 GUARDAR BACKUP EN SESSION STATE
        st.session_state.df_backup = df.copy()
        
        # --- 🔹 1. SGP CSF (Salarios + Parafiscales) ---
        filtro_csf = (df["Codigo"].isin(["2-100-I002", "1-204-I012"]) & 
            (ultimos_cuatro.between(1001, 1013) |
             ultimos_cuatro.between(2020, 2031) |
             ultimos_cuatro.between(3033, 3051))
        )    
        
        csf = {
            "INICIAL": df.loc[filtro_csf, "INICIAL"].sum(),
            "DISPONIBLE": df.loc[filtro_csf, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_csf, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_csf, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_csf, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_csf, "RECURSOS SIN EJECUTAR"].sum()
        }
         
        filtro_FOMAG_Empleado = (df["Codigo"].isin(["2-100-I002", "2-100-I001"]) & (ultimos_cuatro.isin([1017, 3058, 3059])))
         
        FOMAG_Empleado = {
            "INICIAL": df.loc[filtro_FOMAG_Empleado, "INICIAL"].sum(),
            "DISPONIBLE": df.loc[filtro_FOMAG_Empleado, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_FOMAG_Empleado, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_FOMAG_Empleado, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_FOMAG_Empleado, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_FOMAG_Empleado, "RECURSOS SIN EJECUTAR"].sum()
        } 
        
        filtro_FOMAG_SSF_Patron = ((df["Codigo"] == "2-100-I001") & (ultimos_cuatro.isin([1018,1019, 3060, 3061, 3062, 3063])))
         
        FOMAG_SSF_Patron = {
            "INICIAL": df.loc[filtro_FOMAG_SSF_Patron, "INICIAL"].sum(),
            "DISPONIBLE": df.loc[filtro_FOMAG_SSF_Patron, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_FOMAG_SSF_Patron, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_FOMAG_SSF_Patron, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_FOMAG_SSF_Patron, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_FOMAG_SSF_Patron, "RECURSOS SIN EJECUTAR"].sum()
        }
         
        filtro_FOMAG_CSF = ((df["Codigo"] == "2-100-I002") & (ultimos_cuatro.isin([3052, 3054])))
         
        FOMAG_CSF = {
            "INICIAL": df.loc[filtro_FOMAG_CSF, "INICIAL"].sum(),
            "DISPONIBLE": df.loc[filtro_FOMAG_CSF, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_FOMAG_CSF, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_FOMAG_CSF, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_FOMAG_CSF, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_FOMAG_CSF, "RECURSOS SIN EJECUTAR"].sum()
        }
         
        filtro_TOTAL_DOC_SGP = (df["Codigo"].isin(["2-100-I002", "2-100-I001", "1-204-I012"]) 
            & ultimos_cuatro.between(1001, 3063))
         
        TOTAL_DOC_SGP = {
            "INICIAL": df.loc[filtro_TOTAL_DOC_SGP, "INICIAL"].sum(),
            "DISPONIBLE": df.loc[filtro_TOTAL_DOC_SGP, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_TOTAL_DOC_SGP, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_TOTAL_DOC_SGP, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_TOTAL_DOC_SGP, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_TOTAL_DOC_SGP, "RECURSOS SIN EJECUTAR"].sum()
        }
         
        filtro_ADTIVOS_SGP = (df["Codigo"].isin(["2-100-I002", "2-100-I026"]) & (ultimos_cuatro.isin([3065, 3066, 3067, 3068, 3069, 3070, 3071, 3072, 3073, 3074, 3075, 3076, 3077, 3078, 3079, 3080, 3081, 3082, 3083, 3084, 3085, 3086])))
         
        ADTIVOS_SGP = {
            "INICIAL": df.loc[filtro_ADTIVOS_SGP, "DISPONIBLE"].sum(),
            "DISPONIBLE": df.loc[filtro_ADTIVOS_SGP, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_ADTIVOS_SGP, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_ADTIVOS_SGP, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_ADTIVOS_SGP, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_ADTIVOS_SGP, "RECURSOS SIN EJECUTAR"].sum()
        }
         
        filtro_TOTAL_SGP_P8033 = (df["Codigo"].isin(["2-100-I002", "2-100-I001", "2-100-I026", "1-204-I012"]) 
            & ultimos_cuatro.between(1001, 3086))
         
        TOTAL_SGP_P8033 = {
            "INICIAL": df.loc[filtro_TOTAL_SGP_P8033, "INICIAL"].sum(),
            "DISPONIBLE": df.loc[filtro_TOTAL_SGP_P8033, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_TOTAL_SGP_P8033, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_TOTAL_SGP_P8033, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_TOTAL_SGP_P8033, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_TOTAL_SGP_P8033, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_DOC_RP = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.between(1001, 3063)))

        DOC_RP = {
            "INICIAL": df.loc[filtro_DOC_RP, "INICIAL"].sum(),
            "DISPONIBLE": df.loc[filtro_DOC_RP, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_DOC_RP, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_DOC_RP, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_DOC_RP, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_DOC_RP, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_ADTIVOS_RP = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.between(3065, 3086)))

        ADTIVOS_RP = {
            "INICIAL": df.loc[filtro_ADTIVOS_RP, "INICIAL"].sum(),
            "DISPONIBLE": df.loc[filtro_ADTIVOS_RP, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_ADTIVOS_RP, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_ADTIVOS_RP, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_ADTIVOS_RP, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_ADTIVOS_RP, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_SENTENCIAS = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro == 3064))

        SENTENCIAS = {
            "INICIAL": df.loc[filtro_SENTENCIAS, "INICIAL"].sum(),
            "DISPONIBLE": df.loc[filtro_SENTENCIAS, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_SENTENCIAS, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_SENTENCIAS, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_SENTENCIAS, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_SENTENCIAS, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_TOTAL_RP_P8033 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.between(1001, 3086)))

        TOTAL_RP_P8033 = {
            "INICIAL": df.loc[filtro_TOTAL_RP_P8033, "INICIAL"].sum(),
            "DISPONIBLE": df.loc[filtro_TOTAL_RP_P8033, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_TOTAL_RP_P8033, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_TOTAL_RP_P8033, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_TOTAL_RP_P8033, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_TOTAL_RP_P8033, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_PENSIONADOS = ((df["Codigo"] == "2-100-I017") & (ultimos_cuatro == 3087))

        PENSIONADOS = {
            "INICIAL": df.loc[filtro_PENSIONADOS, "INICIAL"].sum(),
            "DISPONIBLE": df.loc[filtro_PENSIONADOS, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_PENSIONADOS, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_PENSIONADOS, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_PENSIONADOS, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_PENSIONADOS, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_TOTAL_GENERAL = (df["Codigo"].isin(["2-100-I002", "2-100-I001","1-100-F001", "2-100-I026", "2-100-I017", "1-204-I012"]) 
            & ultimos_cuatro.between(1001, 3087))
         
        TOTAL_GENERAL = {
            "INICIAL": df.loc[filtro_TOTAL_GENERAL, "INICIAL"].sum(),
            "DISPONIBLE": df.loc[filtro_TOTAL_GENERAL, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_TOTAL_GENERAL, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_TOTAL_GENERAL, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_TOTAL_GENERAL, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_TOTAL_GENERAL, "RECURSOS SIN EJECUTAR"].sum() 
        }

        # --- 🔹 Crear tabla resumen ---
        resumen_principal = pd.DataFrame(
            [
                csf,                                    
                FOMAG_Empleado, FOMAG_SSF_Patron, FOMAG_CSF, TOTAL_DOC_SGP,
                ADTIVOS_SGP, TOTAL_SGP_P8033, DOC_RP, ADTIVOS_RP, SENTENCIAS,
                TOTAL_RP_P8033, PENSIONADOS, TOTAL_GENERAL                   
            ],
            index=[
                "SGP CSF (Salarios + Parafiscales)", "FOMAG_Empleado", "FOMAG_SSF_Patron", 
                "FOMAG_CSF", "TOTAL_DOC_SGP", "ADTIVOS_SGP", "TOTAL_SGP_P8033", "DOC_RP", "ADTIVOS_RP", "SENTENCIAS",
                "TOTAL_RP_P8033", "PENSIONADOS", "TOTAL_GENERAL"
            ]
        )

        resumen = resumen_principal.copy()
        
        return df, resumen
         
    except Exception as e:
        st.error(f"❌ Error en procesar_datos_sgp: {str(e)}")
        import traceback
        st.error(f"Detalle: {traceback.format_exc()}")
        return None, None


# =============================================================================
# FUNCIÓN DE PROCESAMIENTO DOCENTES RP PRINCIPAL 
# =============================================================================
def procesar_datos_RP_principal(fuerza_actualizacion=False):
    """Función específica para procesar datos de RECURSOS PROPIOS"""
    # Cargar datos
    df = cargar_datos_originales(_fuerza_actualizacion=fuerza_actualizacion)
    
    if df is None:
        return None, None
    
    try:        
        # 🔹 CALCULAR 'ultimos_cuatro' LOCALMENTE (IMPORTANTE)
        ultimos_cuatro = pd.to_numeric(
            df["Codigo_O"].astype(str).str[-4:], 
            errors="coerce"
        )
        
        # 🔹 GUARDAR BACKUP EN SESSION STATE
        st.session_state.df_backup = df.copy()
        
        # --- 🔹 1. DEFINIR FILTROS (CORREGIDO el error de sintaxis) ---
        filtro_SUELDO_BASICO = (
            (df["Codigo"] == "1-100-F001") & 
            (ultimos_cuatro.isin([3033, 3034, 3035, 2020, 1001, 1002]))  # ¡CORREGIDO: isin() con paréntesis!
        )    
        
        SUELDO_BASICO = {
            "DISPONIBLE": df.loc[filtro_SUELDO_BASICO, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_SUELDO_BASICO, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_SUELDO_BASICO, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_SUELDO_BASICO, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_SUELDO_BASICO, "RECURSOS SIN EJECUTAR"].sum()
        }
         
        filtro_horas_extras = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3036, 2021, 1003])))
         
        HORAS_EXTRAS = {
            "DISPONIBLE": df.loc[filtro_horas_extras, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_horas_extras, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_horas_extras, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_horas_extras, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_horas_extras, "RECURSOS SIN EJECUTAR"].sum()
        } 
        
        filtro_prima_servicios = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3037, 2022, 1004])))
         
        PRIMA_SERVICIOS = {
            "DISPONIBLE": df.loc[filtro_prima_servicios, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_prima_servicios, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_prima_servicios, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_prima_servicios, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_prima_servicios, "RECURSOS SIN EJECUTAR"].sum()
        }
         
        filtro_prima_vacaciones = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3038, 2023, 1005])))
         
        PRIMA_VACACIONES = {
            "DISPONIBLE": df.loc[filtro_prima_vacaciones, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_prima_vacaciones, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_prima_vacaciones, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_prima_vacaciones, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_prima_vacaciones, "RECURSOS SIN EJECUTAR"].sum()
        }
         
        filtro_prima_navidad = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3039, 2024, 1006])))
         
        PRIMA_NAVIDAD = {
            "DISPONIBLE": df.loc[filtro_prima_navidad, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_prima_navidad, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_prima_navidad, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_prima_navidad, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_prima_navidad, "RECURSOS SIN EJECUTAR"].sum()
        }
         
        filtro_sub_alimentacion = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3040, 1007])))
         
        SUB_ALIMENTACION = {
            "DISPONIBLE": df.loc[filtro_sub_alimentacion, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_sub_alimentacion, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_sub_alimentacion, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_sub_alimentacion, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_sub_alimentacion, "RECURSOS SIN EJECUTAR"].sum()
        }
         
        filtro_aux_transporte = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3041, 1008])))
         
        AUX_TRANSPORTE = {
            "DISPONIBLE": df.loc[filtro_aux_transporte, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_aux_transporte, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_aux_transporte, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_aux_transporte, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_aux_transporte, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_SUELDOS = (
        (df["Codigo"] == "1-100-F001") &
        (
        ultimos_cuatro.between(1001, 1008) |
        ultimos_cuatro.between(2020, 2024) |
        ultimos_cuatro.between(3033, 3041)
        )
         )

        SUELDOS = {
            "DISPONIBLE": df.loc[filtro_SUELDOS, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_SUELDOS, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_SUELDOS, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_SUELDOS, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_SUELDOS, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_compensar = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3042, 3043, 2025, 1009])))

        COMPENSAR = {
            "DISPONIBLE": df.loc[filtro_compensar, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_compensar, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_compensar, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_compensar, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_compensar, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_icbf = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3044, 3045, 2026, 1010])))

        ICBF = {
            "DISPONIBLE": df.loc[filtro_icbf, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_icbf, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_icbf, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_icbf, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_icbf, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_escuelas_tecnicas = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3046, 3047, 2027, 1011])))

        ESCUELAS_TECNICAS = {
            "DISPONIBLE": df.loc[filtro_escuelas_tecnicas, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_escuelas_tecnicas, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_escuelas_tecnicas, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_escuelas_tecnicas, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_escuelas_tecnicas, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_sena = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3048, 3049, 2028, 1012])))

        SENA = {
            "DISPONIBLE": df.loc[filtro_sena, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_sena, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_sena, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_sena, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_sena, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_esap = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3050, 3051, 2029, 1013])))

        ESAP = {
            "DISPONIBLE": df.loc[filtro_esap, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_esap, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_esap, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_esap, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_esap, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_parafiscales = ((df["Codigo"] == "1-100-F001") & 
            (ultimos_cuatro.between(3042, 3051) |
             ultimos_cuatro.between(2025, 2029) |
             ultimos_cuatro.between(1009, 1013)))

        PARAFISCALES = {
            "DISPONIBLE": df.loc[filtro_parafiscales, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_parafiscales, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_parafiscales, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_parafiscales, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_parafiscales, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_salud = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3052, 3053, 2030, 1014])))

        SALUD = {
            "DISPONIBLE": df.loc[filtro_salud, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_salud, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_salud, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_salud, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_salud, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_pension = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3054, 3055, 2031, 1015])))

        PENSION = {
            "DISPONIBLE": df.loc[filtro_pension, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_pension, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_pension, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_pension, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_pension, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_cesantias = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3056, 3057, 2032, 1016])))

        CESANTIAS = {
            "DISPONIBLE": df.loc[filtro_cesantias, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_cesantias, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_cesantias, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_cesantias, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_cesantias, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_fomag = ((df["Codigo"] == "1-100-F001") & 
            (ultimos_cuatro.between(3052, 3057) |
             ultimos_cuatro.between(2030, 2032) |
             ultimos_cuatro.between(1014, 1016)))

        FOMAG = {
            "DISPONIBLE": df.loc[filtro_fomag, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_fomag, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_fomag, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_fomag, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_fomag, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_total_doc_rp = ((df["Codigo"] == "1-100-F001") & 
            (ultimos_cuatro.between(1001, 3057)))

        TOTAL_DOC_RP = {
            "DISPONIBLE": df.loc[filtro_total_doc_rp, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_total_doc_rp, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_total_doc_rp, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_total_doc_rp, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_total_doc_rp, "RECURSOS SIN EJECUTAR"].sum() 
        }

        # --- 🔹 Crear tabla resumen ---
        resumen_principal = pd.DataFrame(
            [
                SUELDO_BASICO, 
                HORAS_EXTRAS, 
                PRIMA_SERVICIOS, 
                PRIMA_VACACIONES, 
                PRIMA_NAVIDAD, 
                SUB_ALIMENTACION, 
                AUX_TRANSPORTE,
                SUELDOS,
                COMPENSAR,
                ICBF,
                ESCUELAS_TECNICAS,
                SENA,
                ESAP,
                PARAFISCALES,
                SALUD,
                PENSION, 
                CESANTIAS, 
                FOMAG, TOTAL_DOC_RP
            ],
            index=[
                "SUELDO BÁSICO", 
                "HORAS EXTRAS", 
                "PRIMA DE SERVICIOS", 
                "PRIMA DE VACACIONES", 
                "PRIMA DE NAVIDAD",
                "SUBSIDIO DE ALIMENTACIÓN", 
                "AUXILIO DE TRANSPORTE",
                "SUELDOS",
                "COMPENSAR",
                "ICBF",
                "ESCUELAS_TECNICAS",
                "SENA",
                "ESAP",
                "PARAFISCALES",
                "SALUD", 
                "PENSION", 
                "CESANTIAS", 
                "FOMAG", "TOTAL_DOC_RP"
            ]
        )

        resumen = resumen_principal.copy()
        
        return df, resumen
         
    except Exception as e:
        st.error(f"❌ Error en procesar_datos_RP_principal: {str(e)}")
        import traceback
        st.error(f"Detalle: {traceback.format_exc()}")
        return None, None

# =============================================================================
# FUNCIÓN DE PROCESAMIENTO DOCENTES RP PRIMERA INFANCIA
# =============================================================================

def procesar_datos_RP_primera_infancia(fuerza_actualizacion=False):
    """Función específica para procesar datos de RECURSOS PROPIOS"""
    # Cargar datos
    df = cargar_datos_originales(_fuerza_actualizacion=fuerza_actualizacion)
    
    if df is None:
        return None, None
    
    try:        
        # 🔹 CALCULAR 'ultimos_dos' LOCALMENTE (IMPORTANTE)
        ultimos_cuatro = pd.to_numeric(
            df["Codigo_O"].astype(str).str[-4:], 
            errors="coerce"
        )
        
        # 🔹 GUARDAR BACKUP EN SESSION STATE
        st.session_state.df_backup = df.copy()
        
        # --- 🔹 1. DEFINIR FILTROS 
        filtro_O2301172201202401690307101001 = (
            (df["Codigo"] == "1-100-F001") & 
            (ultimos_cuatro.isin([1001])) 
        )    
        
        O2301172201202401690307101001 = {
            "NOMBRE": "Pago de Personal Docente primera infancia",
            "CONCEPTO": "O231010100101 Sueldo básico",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307101001, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307101001, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307101001, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307101001, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307101001, "RECURSOS SIN EJECUTAR"].sum()
        }
         
        filtro_O2301172201202401690307101002 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([1002])))
         
        O2301172201202401690307101002 = {
            "NOMBRE": "Pago de Ascensos en escalafon del Personal",
            "CONCEPTO": "O231010100101 Sueldo básico",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307101002, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307101002, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307101002, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307101002, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307101002, "RECURSOS SIN EJECUTAR"].sum()
        } 
        
        filtro_O2301172201202401690307101003 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([1003])))
         
        O2301172201202401690307101003 = {
            "NOMBRE": "Pago de horas extras del personal docente",
            "CONCEPTO": "O231010100102 Horas extras, dominicales, festivos y recargos",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307101003, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307101003, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307101003, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307101003, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307101003, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_O2301172201202401690307101004 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([1004])))
         
        O2301172201202401690307101004 = {
            "NOMBRE": "Pago de Personal Docente prima de servicio",
            "CONCEPTO": "O231010100106 Prima de servicio",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307101004, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307101004, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307101004, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307101004, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307101004, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307101005 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([1005])))
         
        O2301172201202401690307101005 = {
            "NOMBRE": "Pago de Personal Docente prima de vacaciones",
            "CONCEPTO": "O23101010010802 Prima de vacaciones",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307101005, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307101005, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307101005, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307101005, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307101005, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307101006 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([1006])))
         
        O2301172201202401690307101006 = {
            "NOMBRE": "Pago de Personal Docente prima de navidad",
            "CONCEPTO": "O23101010010801 Prima de navidad",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307101006, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307101006, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307101006, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307101006, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307101006, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307101007 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([1007])))
         
        O2301172201202401690307101007 = {
            "NOMBRE": "Pago de Personal Docente subsidio de alimentacion",
            "CONCEPTO": "O231010100104 Subsidio de alimentación",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307101007, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307101007, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307101007, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307101007, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307101007, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307101008 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([1008])))
         
        O2301172201202401690307101008 = {
            "NOMBRE": "Pago Auxili de transporte personal docente",
            "CONCEPTO": "O231010100105 Auxilio de Transporte",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307101008, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307101008, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307101008, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307101008, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307101008, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_sueldos = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([1001, 1002, 1003, 1004, 1005, 1006, 1007, 1008])))
         
        SUELDOS = {
            "NOMBRE": "---",
            "CONCEPTO": "---",
            "DISPONIBLE": df.loc[filtro_sueldos, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_sueldos, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_sueldos, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_sueldos, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_sueldos, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307101009 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([1009])))
         
        O2301172201202401690307101009 = {
            "NOMBRE": "Pago de Aportes para las Cajas de Compen",
            "CONCEPTO": "O231010200401 Compensar",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307101009, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307101009, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307101009, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307101009, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307101009, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307101010 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([1010])))
         
        O2301172201202401690307101010 = {
            "NOMBRE": "Pago de Aportes para el ICBF personal docente",
            "CONCEPTO": "O2310102006 Aportes al ICBF",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307101010, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307101010, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307101010, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307101010, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307101010, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307101011 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([1011])))
         
        O2301172201202401690307101011 = {
            "NOMBRE": "Pago de Aportes para Institutos Tecnicos",
            "CONCEPTO": "O2310102009 Aportes a escuelas industriales e institutos técnicos",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307101011, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307101011, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307101011, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307101011, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307101011, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307101012 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([1012])))
         
        O2301172201202401690307101012 = {
            "NOMBRE": "Pago de Aportes para el SENA personal docentes",
            "CONCEPTO": "O2310102007 Aportes al SENA",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307101012, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307101012, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307101012, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307101012, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307101012, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307101013 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([1013])))
         
        O2301172201202401690307101013 = {
            "NOMBRE": "Pago de Aportes para la ESAP personal docente",
            "CONCEPTO": "O2310102008 Aportes a la ESAP",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307101013, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307101013, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307101013, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307101013, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307101013, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_parafiscales = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([1009, 1010, 1011, 1012, 1013])))
         
        PARAFISCALES = {
            "NOMBRE": "---",
            "CONCEPTO": "---",
            "DISPONIBLE": df.loc[filtro_parafiscales, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_parafiscales, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_parafiscales, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_parafiscales, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_parafiscales, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307101014 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([1014])))
         
        O2301172201202401690307101014 = {
            "NOMBRE": "Pago de Aportes para Salud del personal",
            "CONCEPTO": "O231010200201 Pago de Aportes para Salud del personal",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307101014, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307101014, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307101014, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307101014, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307101014, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307101015 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([1015])))
         
        O2301172201202401690307101015 = {
            "NOMBRE": "Pago de Aportes para Pension del persona",
            "CONCEPTO": "O231010200101 Pago de Aportes para Pension del persona",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307101015, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307101015, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307101015, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307101015, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307101015, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307101016 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([1016])))
         
        O2301172201202401690307101016 = {
            "NOMBRE": "Pago de Aportes para Cesantias del personal",
            "CONCEPTO": "O231010200301 Aportes de cesantías a fondos públicos",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307101016, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307101016, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307101016, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307101016, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307101016, "RECURSOS SIN EJECUTAR"].sum()
        }
        filtro_fomag = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([1014, 1015, 1016])))
         
        FOMAG = {
            "NOMBRE": "---",
            "CONCEPTO": "---",
            "DISPONIBLE": df.loc[filtro_fomag, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_fomag, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_fomag, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_fomag, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_fomag, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_total_primera_infancia = ((df["Codigo"] == "1-100-F001") & 
            (ultimos_cuatro.between(1001, 1016)))
         
        TOTAL_PRIMERA_INFANCIA = {
            "NOMBRE": "---",
            "CONCEPTO": "---",
            "DISPONIBLE": df.loc[filtro_total_primera_infancia, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_total_primera_infancia, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_total_primera_infancia, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_total_primera_infancia, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_total_primera_infancia, "RECURSOS SIN EJECUTAR"].sum()
        }
        

        # --- 🔹 Crear tabla resumen ---
        resumen_principal = pd.DataFrame(
            [
            O2301172201202401690307101001, O2301172201202401690307101002, O2301172201202401690307101003,
            O2301172201202401690307101004, O2301172201202401690307101005, O2301172201202401690307101006,
            O2301172201202401690307101007, O2301172201202401690307101008, SUELDOS, O2301172201202401690307101009,
            O2301172201202401690307101010, O2301172201202401690307101011, O2301172201202401690307101012,
            O2301172201202401690307101013, PARAFISCALES, O2301172201202401690307101014, O2301172201202401690307101015,
            O2301172201202401690307101016, FOMAG, TOTAL_PRIMERA_INFANCIA      
            ],
            index=[
                "O2301172201202401690307101001", "O2301172201202401690307101002", "O2301172201202401690307101003",
                "O2301172201202401690307101004", "O2301172201202401690307101005", "O2301172201202401690307101006",
                "O2301172201202401690307101007", "O2301172201202401690307101008", "SUELDOS", "O2301172201202401690307101009",
                "O2301172201202401690307101010", "O2301172201202401690307101011", "O2301172201202401690307101012",
                "O2301172201202401690307101013", "PARAFISCALES", "O2301172201202401690307101014", "O2301172201202401690307101015",
                "O2301172201202401690307101016", "FOMAG", "TOTAL_PRIMERA_INFANCIA"
            ]
        )

        resumen = resumen_principal.copy()
        
        return df, resumen
         
    except Exception as e:
        st.error(f"❌ Error en procesar_datos_RP_principal: {str(e)}")
        import traceback
        st.error(f"Detalle: {traceback.format_exc()}")
        return None, None


# =============================================================================
# FUNCIÓN DE PROCESAMIENTO DOCENTES RP ORIENTADORES
# =============================================================================

def procesar_datos_RP_orientadores(fuerza_actualizacion=False):
    """Función específica para procesar datos de RECURSOS PROPIOS"""
    # Cargar datos
    df = cargar_datos_originales(_fuerza_actualizacion=fuerza_actualizacion)
    
    if df is None:
        return None, None
    
    try:        
        # 🔹 CALCULAR 'ultimos_dos' LOCALMENTE (IMPORTANTE)
        ultimos_cuatro = pd.to_numeric(
            df["Codigo_O"].astype(str).str[-4:], 
            errors="coerce"
        )
        
        # 🔹 GUARDAR BACKUP EN SESSION STATE
        st.session_state.df_backup = df.copy()
        
        # --- 🔹 1. DEFINIR FILTROS 
        filtro_O2301172201202401690307102020 = (
            (df["Codigo"] == "1-100-F001") & 
            (ultimos_cuatro.isin([2020])) 
        )    
        
        O2301172201202401690307102020 = {
            "NOMBRE": "Pago de Personal Docente orientadores",
            "CONCEPTO": "O231010100101 Sueldo básico",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307102020, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307102020, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307102020, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307102020, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307102020, "RECURSOS SIN EJECUTAR"].sum()
        }
         
        filtro_O2301172201202401690307102021 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([2021])))
         
        O2301172201202401690307102021 = {
            "NOMBRE": "Pago de horas extras del personal docent",
            "CONCEPTO": "O231010100102 Horas extras, dominicales, festivos y recargos",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307102021, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307102021, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307102021, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307102021, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307102021, "RECURSOS SIN EJECUTAR"].sum()
        } 
        
        filtro_O2301172201202401690307102022 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([2022])))
         
        O2301172201202401690307102022 = {
            "NOMBRE": "Pago de Personal Docente prima de servicio",
            "CONCEPTO": "O231010100106 Prima de servicio",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307102022, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307102022, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307102022, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307102022, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307102022, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_O2301172201202401690307102023 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([2023])))
         
        O2301172201202401690307102023 = {
            "NOMBRE": "Pago de Personal Docente prima de vacaciones",
            "CONCEPTO": "O23101010010802 Prima de vacaciones",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307102023, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307102023, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307102023, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307102023, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307102023, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307102024 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([2024])))
         
        O2301172201202401690307102024 = {
            "NOMBRE": "Pago de Personal Docente prima de navidad",
            "CONCEPTO": "O23101010010801 Prima de navidad",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307102024, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307102024, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307102024, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307102024, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307102024, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_sueldos_or = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([2020, 2021, 2022, 2023, 2024])))
         
        SUELDOS_ORIENTADORES = {
            "NOMBRE": "---",
            "CONCEPTO": "---",
            "DISPONIBLE": df.loc[filtro_sueldos_or, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_sueldos_or, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_sueldos_or, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_sueldos_or, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_sueldos_or, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307102025 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([2025])))
         
        O2301172201202401690307102025 = {
            "NOMBRE": "Pago de Aportes para las Cajas de Compensacion",
            "CONCEPTO": "O231010200401 Compensar",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307102025, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307102025, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307102025, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307102025, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307102025, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307102026 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([2026])))
         
        O2301172201202401690307102026 = {
            "NOMBRE": "Pago de Aportes para el ICBF personal docente",
            "CONCEPTO": "O2310102006 Aportes al ICBF",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307102026, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307102026, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307102026, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307102026, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307102026, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307102027 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([2027])))
         
        O2301172201202401690307102027 = {
            "NOMBRE": "Pago de Aportes para Institutos Tecnicos",
            "CONCEPTO": "O2310102009 Aportes a escuelas industriales e institutos técnicos",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307102027, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307102027, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307102027, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307102027, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307102027, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307102028 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([2028])))
         
        O2301172201202401690307102028 = {
            "NOMBRE": "Pago de Aportes para el SENA personal docente",
            "CONCEPTO": "O2310102007 Aportes al SENA",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307102028, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307102028, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307102028, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307102028, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307102028, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307102029 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([2029])))
         
        O2301172201202401690307102029 = {
            "NOMBRE": "Pago de Aportes para la ESAP personal docente",
            "CONCEPTO": "O2310102008 Aportes a la ESAP",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307102029, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307102029, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307102029, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307102029, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307102029, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_parafiscales_or = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([2025, 2026, 2027, 2028, 2029])))
         
        PARAFISCALES_ORIENTADORES = {
            "NOMBRE": "---",
            "CONCEPTO": "---",
            "DISPONIBLE": df.loc[filtro_parafiscales_or, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_parafiscales_or, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_parafiscales_or, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_parafiscales_or, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_parafiscales_or, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307102030 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([2030])))
         
        O2301172201202401690307102030 = {
            "NOMBRE": "Pago de Aportes para Salud del personal",
            "CONCEPTO": "O231010200201 Aportes a la seguridad social en salud pública",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307102030, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307102030, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307102030, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307102030, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307102030, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307102031 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([2031])))
         
        O2301172201202401690307102031 = {
            "NOMBRE": "Pago de Aportes para Pension del personal",
            "CONCEPTO": "O231010200101 Aportes a la seguridad social en pensiones públicas",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307102031, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307102031, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307102031, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307102031, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307102031, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307102032 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([2032])))
         
        O2301172201202401690307102032 = {
            "NOMBRE": "Pago de Aportes para Cesantias del personal",
            "CONCEPTO": "O231010200301 Aportes de cesantías a fondos públicos",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307102032, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307102032, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307102032, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307102032, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307102032, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_fomag_or = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([2030, 2031, 2032])))
         
        FOMAG_ORIENTADORES = {
            "NOMBRE": "---",
            "CONCEPTO": "---",
            "DISPONIBLE": df.loc[filtro_fomag_or, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_fomag_or, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_fomag_or, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_fomag_or, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_fomag_or, "RECURSOS SIN EJECUTAR"].sum()
        }
        filtro_TOTAL_DOC_ORIENTADORES = ((df["Codigo"] == "1-100-F001") & 
            (ultimos_cuatro.between(2020, 2032)))
         
        TOTAL_DOC_ORIENTADORES = {
            "NOMBRE": "---",
            "CONCEPTO": "---",
            "DISPONIBLE": df.loc[filtro_TOTAL_DOC_ORIENTADORES, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_TOTAL_DOC_ORIENTADORES, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_TOTAL_DOC_ORIENTADORES, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_TOTAL_DOC_ORIENTADORES, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_TOTAL_DOC_ORIENTADORES, "RECURSOS SIN EJECUTAR"].sum()
        }

        # --- 🔹 Crear tabla resumen ---
        resumen_principal = pd.DataFrame(
            [
            O2301172201202401690307102020, O2301172201202401690307102021, O2301172201202401690307102022,
            O2301172201202401690307102023, O2301172201202401690307102024, SUELDOS_ORIENTADORES, 
            O2301172201202401690307102025, 
            O2301172201202401690307102026, O2301172201202401690307102027, O2301172201202401690307102028,
            O2301172201202401690307102029, PARAFISCALES_ORIENTADORES, 
            O2301172201202401690307102030, O2301172201202401690307102031, 
            O2301172201202401690307102032, FOMAG_ORIENTADORES, TOTAL_DOC_ORIENTADORES    
            ],
            index=[
                "O2301172201202401690307102020", "O2301172201202401690307102021", "O2301172201202401690307102022",
            "O2301172201202401690307102023", "O2301172201202401690307102024", "SUELDOS_ORIENTADORES", 
            "O2301172201202401690307102025", 
            "O2301172201202401690307102026", "O2301172201202401690307102027", "O2301172201202401690307102028",
            "O2301172201202401690307102029", "PARAFISCALES_ORIENTADORES", 
            "O2301172201202401690307102030", "O2301172201202401690307102031", 
            "O2301172201202401690307102032", "FOMAG_ORIENTADORES", "TOTAL_DOC_ORIENTADORES"
            ]
        )

        resumen = resumen_principal.copy()
        
        return df, resumen
         
    except Exception as e:
        st.error(f"❌ Error en procesar_datos_RP_principal: {str(e)}")
        import traceback
        st.error(f"Detalle: {traceback.format_exc()}")
        return None, None

# =============================================================================
# FUNCIÓN DE PROCESAMIENTO DOCENTES RP primaria,basica,media
# =============================================================================

def procesar_datos_RP_primaria_basica_media(fuerza_actualizacion=False):
    """Función específica para procesar datos de RECURSOS PROPIOS"""
    # Cargar datos
    df = cargar_datos_originales(_fuerza_actualizacion=fuerza_actualizacion)
    
    if df is None:
        return None, None
    
    try:        
        # 🔹 CALCULAR 'ultimos_dos' LOCALMENTE (IMPORTANTE)
        ultimos_cuatro = pd.to_numeric(
            df["Codigo_O"].astype(str).str[-4:], 
            errors="coerce"
        )
        
        # 🔹 GUARDAR BACKUP EN SESSION STATE
        st.session_state.df_backup = df.copy()
        
        # --- 🔹 1. DEFINIR FILTROS 
        filtro_O2301172201202401690307103033 = (
            (df["Codigo"] == "1-100-F001") & 
            (ultimos_cuatro.isin([3033])) 
        )    
        
        O2301172201202401690307103033 = {
            "NOMBRE": "Pago de Personal Docente",
            "CONCEPTO": "O231010100101 Sueldo básico",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103033, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103033, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103033, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103033, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103033, "RECURSOS SIN EJECUTAR"].sum()
        }
         
        filtro_O2301172201202401690307103034 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3034])))
         
        O2301172201202401690307103034 = {
            "NOMBRE": "Pago de Personal Directivo Docente",
            "CONCEPTO": "O231010100101 Sueldo básico",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103034, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103034, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103034, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103034, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103034, "RECURSOS SIN EJECUTAR"].sum()
        } 
        
        filtro_O2301172201202401690307103035 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3035])))
         
        O2301172201202401690307103035 = {
            "NOMBRE": "Pago de Ascensos en escalafon del Personal",
            "CONCEPTO": "O231010100101 Sueldo básico",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103035, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103035, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103035, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103035, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103035, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_O2301172201202401690307103036 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3036])))
         
        O2301172201202401690307103036 = {
            "NOMBRE": "Pago de horas extras del personal docent",
            "CONCEPTO": "O231010100102 Horas extras, dominicales, festivos y recargos",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103036, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103036, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103036, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103036, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103036, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307103037 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3037])))
         
        O2301172201202401690307103037 = {
            "NOMBRE": "Pago de Personal Docente- prima de servicio",
            "CONCEPTO": "O231010100106 Prima de servicio",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103037, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103037, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103037, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103037, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103037, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307103038 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3038])))
         
        O2301172201202401690307103038 = {
            "NOMBRE": "Pago de Personal Docente - prima de vacaciones",
            "CONCEPTO": "O23101010010802 Prima de vacaciones",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103038, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103038, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103038, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103038, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103038, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307103039 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3039])))
         
        O2301172201202401690307103039 = {
            "NOMBRE": "Pago de Personal Docente - prima de navidad",
            "CONCEPTO": "O23101010010801 Prima de navidad",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103039, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103039, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103039, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103039, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103039, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307103040 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3040])))
         
        O2301172201202401690307103040 = {
            "NOMBRE": "Pago de Personal Docente- subsidio de alimentación",
            "CONCEPTO": "O231010100104 Subsidio de alimentación",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103040, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103040, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103040, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103040, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103040, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307103041 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3041])))
         
        O2301172201202401690307103041 = {
            "NOMBRE": "Pago Auxili de transporte personal docente",
            "CONCEPTO": "O231010100105 Auxilio de Transporte",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103041, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103041, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103041, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103041, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103041, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_sueldos_gl = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.between(3033, 3041)))
         
        SUELDOS_PBM = {
            "NOMBRE": "---",
            "CONCEPTO": "---",
            "DISPONIBLE": df.loc[filtro_sueldos_gl, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_sueldos_gl, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_sueldos_gl, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_sueldos_gl, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_sueldos_gl, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307103042 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3042])))
         
        O2301172201202401690307103042 = {
            "NOMBRE": "Pago de Aportes para las Cajas de Compension",
            "CONCEPTO": "O231010200401 Compensar",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103042, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103042, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103042, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103042, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103042, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307103043 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3043])))
         
        O2301172201202401690307103043 = {
            "NOMBRE": "Pago de Aportes para las Cajas de Compensacion",
            "CONCEPTO": "O231010200401 Compensar",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103043, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103043, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103043, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103043, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103043, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307103044 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3044])))
         
        O2301172201202401690307103044 = {
            "NOMBRE": "Pago de Aportes para el ICBF personal do",
            "CONCEPTO": "O2310102006 Aportes al ICBF",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103044, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103044, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103044, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103044, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103044, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307103045 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3045])))
         
        O2301172201202401690307103045 = {
            "NOMBRE": "Pago de Aportes para el ICBF del Personal",
            "CONCEPTO": "O2310102006 Aportes al ICBF",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103045, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103045, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103045, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103045, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103045, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307103046 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3046])))
         
        O2301172201202401690307103046 = {
            "NOMBRE": "Pago de Aportes para Institutos Tecnicos",
            "CONCEPTO": "O2310102009 Aportes a escuelas industriales e institutos técnicos",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103046, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103046, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103046, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103046, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103046, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307103047 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3047])))
         
        O2301172201202401690307103047 = {
            "NOMBRE": "Pago de Aportes para Institutos Tecnicos",
            "CONCEPTO": "O2310102009 Aportes a escuelas industriales e institutos técnicos",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103047, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103047, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103047, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103047, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103047, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307103048 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3048])))
         
        O2301172201202401690307103048 = {
            "NOMBRE": "Pago de Aportes para el SENA del Personal",
            "CONCEPTO": "O2310102007 Aportes al SENA",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103048, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103048, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103048, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103048, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103048, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307103049 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3049])))
         
        O2301172201202401690307103049 = {
            "NOMBRE": "Pago de Aportes para el SENA del Personal",
            "CONCEPTO": "O2310102007 Aportes al SENA",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103049, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103049, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103049, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103049, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103049, "RECURSOS SIN EJECUTAR"].sum()
        }
        filtro_O2301172201202401690307103050 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3050])))
         
        O2301172201202401690307103050 = {
            "NOMBRE": "Pago de Aportes para la ESAP personal docente",
            "CONCEPTO": "O2310102008 Aportes a la ESAP",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103050, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103050, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103050, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103050, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103050, "RECURSOS SIN EJECUTAR"].sum()
        }
        filtro_O2301172201202401690307103051 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3051])))
         
        O2301172201202401690307103051 = {
            "NOMBRE": "Pago de Aportes para la ESAP personal docente",
            "CONCEPTO": "O2310102008 Aportes a la ESAP",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103051, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103051, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103051, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103051, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103051, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_PARAFISCALES_PBM = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.between(3042, 3051)))
         
        PARAFISCALES_PBM = {
            "NOMBRE": "---",
            "CONCEPTO": "---",
            "DISPONIBLE": df.loc[filtro_PARAFISCALES_PBM, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_PARAFISCALES_PBM, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_PARAFISCALES_PBM, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_PARAFISCALES_PBM, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_PARAFISCALES_PBM, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_O2301172201202401690307103052 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3052])))
         
        O2301172201202401690307103052 = {
            "NOMBRE": "Pago de Aportes para Salud del personal",
            "CONCEPTO": "O231010200201 Aportes a la seguridad social en salud pública",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103052, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103052, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103052, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103052, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103052, "RECURSOS SIN EJECUTAR"].sum()
        }
        filtro_O2301172201202401690307103053 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3053])))
         
        O2301172201202401690307103053 = {
            "NOMBRE": "Pago de Aportes para salud del personal",
            "CONCEPTO": "O231010200201 Aportes a la seguridad social en salud pública",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103053, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103053, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103053, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103053, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103053, "RECURSOS SIN EJECUTAR"].sum()
        }
        filtro_O2301172201202401690307103054 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3054])))
         
        O2301172201202401690307103054 = {
            "NOMBRE": "Pago de Aportes para Pension del personal",
            "CONCEPTO": "O231010200101 Aportes a la seguridad social en pensiones públicas",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103054, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103054, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103054, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103054, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103054, "RECURSOS SIN EJECUTAR"].sum()
        }
        filtro_O2301172201202401690307103055 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3055])))
         
        O2301172201202401690307103055 = {
            "NOMBRE": "Pago de Aportes para Pension del personal",
            "CONCEPTO": "O231010200101 Aportes a la seguridad social en pensiones públicas",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103055, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103055, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103055, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103055, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103055, "RECURSOS SIN EJECUTAR"].sum()
        }
        filtro_O2301172201202401690307103056 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3056])))
         
        O2301172201202401690307103056 = {
            "NOMBRE": "Pago de Aportes para Cesantias del personal",
            "CONCEPTO": "O231010200301 Aportes de cesantías a fondos públicos",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103056, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103056, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103056, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103056, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103056, "RECURSOS SIN EJECUTAR"].sum()
        }
        filtro_O2301172201202401690307103057 = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.isin([3057])))
         
        O2301172201202401690307103057 = {
            "NOMBRE": "Pago de Aportes para Cesantias del personal",
            "CONCEPTO": "O231010200301 Aportes de cesantías a fondos públicos",
            "DISPONIBLE": df.loc[filtro_O2301172201202401690307103057, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_O2301172201202401690307103057, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_O2301172201202401690307103057, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_O2301172201202401690307103057, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_O2301172201202401690307103057, "RECURSOS SIN EJECUTAR"].sum()
        }

        filtro_FOMAG_PBM = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.between(3052, 3057)))
         
        FOMAG_PBM = {
            "NOMBRE": "---",
            "CONCEPTO": "---",
            "DISPONIBLE": df.loc[filtro_FOMAG_PBM, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_FOMAG_PBM, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_FOMAG_PBM, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_FOMAG_PBM, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_FOMAG_PBM, "RECURSOS SIN EJECUTAR"].sum()
        }
        filtro_TOTAL_DOC_PBM = ((df["Codigo"] == "1-100-F001") & (ultimos_cuatro.between(3033, 3057)))
         
        TOTAL_DOC_PBM = {
            "NOMBRE": "---",
            "CONCEPTO": "---",
            "DISPONIBLE": df.loc[filtro_TOTAL_DOC_PBM, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_TOTAL_DOC_PBM, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_TOTAL_DOC_PBM, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_TOTAL_DOC_PBM, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_TOTAL_DOC_PBM, "RECURSOS SIN EJECUTAR"].sum()
        }

        # --- 🔹 Crear tabla resumen ---
        resumen_principal = pd.DataFrame(
            [
            O2301172201202401690307103033, O2301172201202401690307103034, O2301172201202401690307103035,
            O2301172201202401690307103036, O2301172201202401690307103037, O2301172201202401690307103038,
            O2301172201202401690307103039, O2301172201202401690307103040, O2301172201202401690307103041,
            SUELDOS_PBM,
            O2301172201202401690307103042, O2301172201202401690307103043, O2301172201202401690307103044, 
            O2301172201202401690307103045, O2301172201202401690307103046, O2301172201202401690307103047,
            O2301172201202401690307103048, O2301172201202401690307103049, O2301172201202401690307103050,
            O2301172201202401690307103051, PARAFISCALES_PBM, 
            O2301172201202401690307103052, O2301172201202401690307103053, 
            O2301172201202401690307103054, O2301172201202401690307103055, O2301172201202401690307103056,
            O2301172201202401690307103057, FOMAG_PBM, TOTAL_DOC_PBM   
            ],
            index=[
            "O2301172201202401690307103033", "O2301172201202401690307103034", "O2301172201202401690307103035",
            "O2301172201202401690307103036", "O2301172201202401690307103037", "O2301172201202401690307103038",
            "O2301172201202401690307103039", "O2301172201202401690307103040", "O2301172201202401690307103041",
            "SUELDOS_PBM",
            "O2301172201202401690307103042", "O2301172201202401690307103043", "O2301172201202401690307103044", 
            "O2301172201202401690307103045", "O2301172201202401690307103046", "O2301172201202401690307103047",
            "O2301172201202401690307103048", "O2301172201202401690307103049", "O2301172201202401690307103050",
            "O2301172201202401690307103051", "PARAFISCALES_PBM", 
            "O2301172201202401690307103052", "O2301172201202401690307103053", 
            "O2301172201202401690307103054", "O2301172201202401690307103055", "O2301172201202401690307103056",
            "O2301172201202401690307103057", "FOMAG_PBM", "TOTAL_DOC_PBM"
            ]
        )

        resumen = resumen_principal.copy()
        
        return df, resumen
         
    except Exception as e:
        st.error(f"❌ Error en procesar_datos_RP_principal: {str(e)}")
        import traceback
        st.error(f"Detalle: {traceback.format_exc()}")
        return None, None

# =============================================================================
# FUNCIÓN DE PROCESAMIENTO DOCENTES SGP PRINCIPAL 
# =============================================================================
def procesar_datos_SGP_principal(fuerza_actualizacion=False):
    """Función específica para procesar datos de SGP"""
    # Cargar datos
    df = cargar_datos_originales(_fuerza_actualizacion=fuerza_actualizacion)
    
    if df is None:
        return None, None
    
    try:        
        # 🔹 CALCULAR 'ultimos_cuatro' LOCALMENTE (IMPORTANTE)
        ultimos_cuatro = pd.to_numeric(
            df["Codigo_O"].astype(str).str[-4:], 
            errors="coerce"
        )
        
        # 🔹 GUARDAR BACKUP EN SESSION STATE
        st.session_state.df_backup = df.copy()
        
        # --- 🔹 1. DEFINIR FILTROS (CORREGIDO el error de sintaxis) ---
        filtro_SUELDO_BASICO = (
            (df["Codigo"] == "2-100-I002") & 
            (ultimos_cuatro.isin([3033, 3034, 3035, 1001, 1002]))  # ¡CORREGIDO: isin() con paréntesis!
        )    
        
        SUELDO_BASICO = {
            "DISPONIBLE": df.loc[filtro_SUELDO_BASICO, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_SUELDO_BASICO, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_SUELDO_BASICO, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_SUELDO_BASICO, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_SUELDO_BASICO, "RECURSOS SIN EJECUTAR"].sum()
        }
         
        filtro_horas_extras = ((df["Codigo"] == "2-100-I002") & (ultimos_cuatro.isin([3036, 1003])))
         
        HORAS_EXTRAS = {
            "DISPONIBLE": df.loc[filtro_horas_extras, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_horas_extras, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_horas_extras, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_horas_extras, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_horas_extras, "RECURSOS SIN EJECUTAR"].sum()
        } 
        
        filtro_prima_servicios = ((df["Codigo"] == "2-100-I002") & (ultimos_cuatro.isin([3037, 1004])))
         
        PRIMA_SERVICIOS = {
            "DISPONIBLE": df.loc[filtro_prima_servicios, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_prima_servicios, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_prima_servicios, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_prima_servicios, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_prima_servicios, "RECURSOS SIN EJECUTAR"].sum()
        }
         
        filtro_prima_vacaciones = ((df["Codigo"] == "2-100-I002") & (ultimos_cuatro.isin([3038, 1005])))
         
        PRIMA_VACACIONES = {
            "DISPONIBLE": df.loc[filtro_prima_vacaciones, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_prima_vacaciones, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_prima_vacaciones, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_prima_vacaciones, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_prima_vacaciones, "RECURSOS SIN EJECUTAR"].sum()
        }
         
        filtro_prima_navidad = ((df["Codigo"] == "2-100-I002") & (ultimos_cuatro.isin([3039, 1006])))
         
        PRIMA_NAVIDAD = {
            "DISPONIBLE": df.loc[filtro_prima_navidad, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_prima_navidad, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_prima_navidad, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_prima_navidad, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_prima_navidad, "RECURSOS SIN EJECUTAR"].sum()
        }
         
        filtro_sub_alimentacion = ((df["Codigo"] == "2-100-I002") & (ultimos_cuatro.isin([3040, 1007])))
         
        SUB_ALIMENTACION = {
            "DISPONIBLE": df.loc[filtro_sub_alimentacion, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_sub_alimentacion, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_sub_alimentacion, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_sub_alimentacion, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_sub_alimentacion, "RECURSOS SIN EJECUTAR"].sum()
        }
         
        filtro_aux_transporte = ((df["Codigo"] == "2-100-I002") & (ultimos_cuatro.isin([3041, 1008])))
         
        AUX_TRANSPORTE = {
            "DISPONIBLE": df.loc[filtro_aux_transporte, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_aux_transporte, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_aux_transporte, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_aux_transporte, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_aux_transporte, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_SUELDOS = (
        (df["Codigo"] == "2-100-I002") &
        (
        ultimos_cuatro.between(1001, 1008) |
        ultimos_cuatro.between(3033, 3041)
        )
         )

        SUELDOS = {
            "DISPONIBLE": df.loc[filtro_SUELDOS, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_SUELDOS, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_SUELDOS, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_SUELDOS, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_SUELDOS, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_compensar = ((df["Codigo"] == "2-100-I002") & (ultimos_cuatro.isin([3042, 3043, 1009])))

        COMPENSAR = {
            "DISPONIBLE": df.loc[filtro_compensar, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_compensar, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_compensar, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_compensar, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_compensar, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_icbf = ((df["Codigo"] == "2-100-I002") & (ultimos_cuatro.isin([3044, 3045, 1010])))

        ICBF = {
            "DISPONIBLE": df.loc[filtro_icbf, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_icbf, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_icbf, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_icbf, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_icbf, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_escuelas_tecnicas = ((df["Codigo"] == "2-100-I002") & (ultimos_cuatro.isin([3046, 3047, 1011])))

        ESCUELAS_TECNICAS = {
            "DISPONIBLE": df.loc[filtro_escuelas_tecnicas, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_escuelas_tecnicas, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_escuelas_tecnicas, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_escuelas_tecnicas, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_escuelas_tecnicas, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_sena = ((df["Codigo"] == "2-100-I002") & (ultimos_cuatro.isin([3048, 3049, 1012])))

        SENA = {
            "DISPONIBLE": df.loc[filtro_sena, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_sena, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_sena, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_sena, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_sena, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_esap = ((df["Codigo"] == "2-100-I002") & (ultimos_cuatro.isin([3050, 3051, 1013])))

        ESAP = {
            "DISPONIBLE": df.loc[filtro_esap, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_esap, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_esap, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_esap, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_esap, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_parafiscales = ((df["Codigo"] == "2-100-I002") & 
            (ultimos_cuatro.between(3042, 3051) 
              |
             ultimos_cuatro.between(1009, 1013)))

        PARAFISCALES = {
            "DISPONIBLE": df.loc[filtro_parafiscales, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_parafiscales, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_parafiscales, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_parafiscales, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_parafiscales, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_fomag_SSF_empleado = ((df["Codigo"] == "2-100-I001") & (ultimos_cuatro.isin([1017, 3058, 3059 ])))

        FOMAG_SSF_EMPLEADO = {
            "DISPONIBLE": df.loc[filtro_fomag_SSF_empleado, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_fomag_SSF_empleado, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_fomag_SSF_empleado, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_fomag_SSF_empleado, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_fomag_SSF_empleado, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_salud_SSF = ((df["Codigo"] == "2-100-I001") & (ultimos_cuatro.isin([1018, 1014, 3052, 3053, 3060, 3061])))

        SALUD_SSF = {
            "DISPONIBLE": df.loc[filtro_salud_SSF, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_salud_SSF, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_salud_SSF, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_salud_SSF, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_salud_SSF, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_pension_ssf = ((df["Codigo"] == "2-100-I001") & (ultimos_cuatro.isin([1015, 3054, 3055])))

        PENSION_SSF = {
            "DISPONIBLE": df.loc[filtro_pension_ssf, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_pension_ssf, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_pension_ssf, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_pension_ssf, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_pension_ssf, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_cesantias_ssf = ((df["Codigo"] == "2-100-I001") & (ultimos_cuatro.isin([1019, 1016, 3056, 3057, 3062, 3063])))

        CESANTIAS_SSF = {
            "DISPONIBLE": df.loc[filtro_cesantias_ssf, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_cesantias_ssf, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_cesantias_ssf, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_cesantias_ssf, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_cesantias_ssf, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_fomag_ssf = ((df["Codigo"] == "2-100-I001") & 
            (ultimos_cuatro.between(3052, 3063)  |
             ultimos_cuatro.between(1014, 1019)))

        FOMAG_ssf = {
            "DISPONIBLE": df.loc[filtro_fomag_ssf, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_fomag_ssf, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_fomag_ssf, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_fomag_ssf, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_fomag_ssf, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_fomag_CSF_empleado = ((df["Codigo"] == "2-100-I002") & (ultimos_cuatro.isin([1017, 3058, 3059 ])))

        FOMAG_CSF_EMPLEADO = {
            "DISPONIBLE": df.loc[filtro_fomag_CSF_empleado, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_fomag_CSF_empleado, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_fomag_CSF_empleado, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_fomag_CSF_empleado, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_fomag_CSF_empleado, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_salud_CSF = ((df["Codigo"].isin(["2-100-I002", "1-200-I038"])) & (ultimos_cuatro.isin([1018, 1014, 3052, 3053, 3060, 3061])))

        SALUD_CSF = {
            "DISPONIBLE": df.loc[filtro_salud_CSF, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_salud_CSF, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_salud_CSF, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_salud_CSF, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_salud_CSF, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_pension_csf = ((df["Codigo"].isin(["2-100-I002", "1-200-I038"])) & (ultimos_cuatro.isin([1015, 3054, 3055])))

        PENSION_CSF = {
            "DISPONIBLE": df.loc[filtro_pension_csf, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_pension_csf, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_pension_csf, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_pension_csf, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_pension_csf, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_cesantias_csf = ((df["Codigo"].isin(["2-100-I002", "1-200-I038"])) & (ultimos_cuatro.isin([1016, 1019, 3056, 3057, 3062, 3063])))

        CESANTIAS_CSF = {
            "DISPONIBLE": df.loc[filtro_cesantias_csf, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_cesantias_csf, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_cesantias_csf, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_cesantias_csf, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_cesantias_csf, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_fomag_csf = ((df["Codigo"].isin(["2-100-I002", "1-200-I038"])) & 
            (ultimos_cuatro.between(3052, 3063)  |
             ultimos_cuatro.between(1014, 1019)))

        FOMAG_csf = {
            "DISPONIBLE": df.loc[filtro_fomag_csf, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_fomag_csf, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_fomag_csf, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_fomag_csf, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_fomag_csf, "RECURSOS SIN EJECUTAR"].sum() 
        }

        filtro_total_doc_sgp = ((df["Codigo"].isin(["2-100-I002", "1-200-I038", "2-100-I001"])) & (ultimos_cuatro.between(1001, 3063)))

        TOTAL_DOC_SGP = {
            "DISPONIBLE": df.loc[filtro_total_doc_sgp, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_total_doc_sgp, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_total_doc_sgp, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_total_doc_sgp, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_total_doc_sgp, "RECURSOS SIN EJECUTAR"].sum() 
        }

        # --- 🔹 Crear tabla resumen ---
        resumen_principal = pd.DataFrame(
            [
                SUELDO_BASICO, 
                HORAS_EXTRAS, 
                PRIMA_SERVICIOS, 
                PRIMA_VACACIONES, 
                PRIMA_NAVIDAD, 
                SUB_ALIMENTACION, 
                AUX_TRANSPORTE,
                SUELDOS,
                COMPENSAR,
                ICBF,
                ESCUELAS_TECNICAS,
                SENA,
                ESAP,
                PARAFISCALES,
                FOMAG_SSF_EMPLEADO,
                SALUD_SSF,
                PENSION_SSF,
                CESANTIAS_SSF,
                FOMAG_ssf,
                FOMAG_CSF_EMPLEADO,
                SALUD_CSF,
                PENSION_CSF,
                CESANTIAS_CSF,
                FOMAG_csf, 
                TOTAL_DOC_SGP
            ],
            index=[
                "SUELDO_BASICO", 
                "HORAS_EXTRAS", 
                "PRIMA_SERVICIOS", 
                "PRIMA_VACACIONES", 
                "PRIMA_NAVIDAD", 
                "SUB_ALIMENTACION", 
                "AUX_TRANSPORTE",
                "SUELDOS",
                "COMPENSAR",
                "ICBF",
                "ESCUELAS_TECNICAS",
                "SENA",
                "ESAP",
                "PARAFISCALES",
                "FOMAG_SSF_EMPLEADO",
                "SALUD_SSF",
                "PENSION_SSF",
                "CESANTIAS_SSF",
                "FOMAG_ssf",
                "FOMAG_CSF_EMPLEADO",
                "SALUD_CSF",
                "PENSION_CSF",
                "CESANTIAS_CSF",
                "FOMAG_csf", 
                "TOTAL_DOC_SGP"
            ]
        )

        resumen = resumen_principal.copy()
        
        return df, resumen
         
    except Exception as e:
        st.error(f"❌ Error en procesar_datos_RP_principal: {str(e)}")
        import traceback
        st.error(f"Detalle: {traceback.format_exc()}")
        return None, None

# =============================================================================
# FUNCIÓN DE PROCESAMIENTO SGP PRIMERA INFANCIA - DETALLADA (con NOMBRE y CONCEPTO)
# =============================================================================

def procesar_datos_SGP_primera_infancia_detallada(fuerza_actualizacion=False):
    """Función específica para procesar datos de SGP - Primera Infancia (códigos 1001-1019) con detalle"""
    df = cargar_datos_originales(_fuerza_actualizacion=fuerza_actualizacion)
    
    if df is None:
        return None, None
    
    try:        
        ultimos_cuatro = pd.to_numeric(
            df["Codigo_O"].astype(str).str[-4:], 
            errors="coerce"
        )
        
        st.session_state.df_backup = df.copy()
        
        # --- 🔹 DEFINIR FILTROS INDIVIDUALES PARA PRIMERA INFANCIA ---
        
        # SUELDOS (1001-1008)
        filtro_1001 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 1001)
        O2301172201202401690307101001 = {
            "NOMBRE": "Pago de Personal Docente primera infancia",
            "CONCEPTO": "O231010100101 Sueldo basico",
            "DISPONIBLE": df.loc[filtro_1001, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_1001, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_1001, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_1001, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_1001, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_1002 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 1002)
        O2301172201202401690307101002 = {
            "NOMBRE": "Pago de Ascensos en escalafon del Personal",
            "CONCEPTO": "O231010100101 Sueldo basico",
            "DISPONIBLE": df.loc[filtro_1002, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_1002, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_1002, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_1002, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_1002, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_1003 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 1003)
        O2301172201202401690307101003 = {
            "NOMBRE": "Pago de horas extras del personal docente",
            "CONCEPTO": "O231010100102 Horas extras, dominicales, festivos y recargos",
            "DISPONIBLE": df.loc[filtro_1003, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_1003, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_1003, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_1003, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_1003, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_1004 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 1004)
        O2301172201202401690307101004 = {
            "NOMBRE": "Pago de Personal Docente prima de servicio",
            "CONCEPTO": "O231010100106 Prima de servicio",
            "DISPONIBLE": df.loc[filtro_1004, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_1004, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_1004, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_1004, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_1004, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_1005 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 1005)
        O2301172201202401690307101005 = {
            "NOMBRE": "Pago de Personal Docente prima de vacaciones",
            "CONCEPTO": "O23101010010802 Prima de vacaciones",
            "DISPONIBLE": df.loc[filtro_1005, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_1005, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_1005, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_1005, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_1005, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_1006 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 1006)
        O2301172201202401690307101006 = {
            "NOMBRE": "Pago de Personal Docente prima de navidad",
            "CONCEPTO": "O23101010010801 Prima de navidad",
            "DISPONIBLE": df.loc[filtro_1006, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_1006, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_1006, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_1006, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_1006, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_1007 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 1007)
        O2301172201202401690307101007 = {
            "NOMBRE": "Pago de Personal Docente subsidio de alimentacion",
            "CONCEPTO": "O231010100104 Subsidio de alimentacion",
            "DISPONIBLE": df.loc[filtro_1007, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_1007, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_1007, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_1007, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_1007, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_1008 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 1008)
        O2301172201202401690307101008 = {
            "NOMBRE": "Pago Auxili de transporte personal docente",
            "CONCEPTO": "O231010100105 Auxilio de Transporte",
            "DISPONIBLE": df.loc[filtro_1008, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_1008, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_1008, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_1008, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_1008, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        # Total SUELDOS (1001-1008)
        filtro_sueldos = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro.between(1001, 1008))
        SUELDOS = {
            "NOMBRE": "---",
            "CONCEPTO": "---",
            "DISPONIBLE": df.loc[filtro_sueldos, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_sueldos, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_sueldos, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_sueldos, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_sueldos, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        # PARAFISCALES (1009-1013)
        filtro_1009 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 1009)
        O2301172201202401690307101009 = {
            "NOMBRE": "Pago de Aportes para las Cajas de Compensacion",
            "CONCEPTO": "O231010200401 Compensar",
            "DISPONIBLE": df.loc[filtro_1009, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_1009, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_1009, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_1009, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_1009, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_1010 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 1010)
        O2301172201202401690307101010 = {
            "NOMBRE": "Pago de Aportes para el ICBF personal docente",
            "CONCEPTO": "O2310102006 Aportes al ICBF",
            "DISPONIBLE": df.loc[filtro_1010, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_1010, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_1010, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_1010, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_1010, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_1011 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 1011)
        O2301172201202401690307101011 = {
            "NOMBRE": "Pago de Aportes para Institutos Tecnicos",
            "CONCEPTO": "O2310102009 Aportes a escuelas industriales e institutos tecnicos",
            "DISPONIBLE": df.loc[filtro_1011, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_1011, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_1011, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_1011, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_1011, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_1012 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 1012)
        O2301172201202401690307101012 = {
            "NOMBRE": "Pago de Aportes para el SENA personal docente",
            "CONCEPTO": "O2310102007 Aportes al SENA",
            "DISPONIBLE": df.loc[filtro_1012, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_1012, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_1012, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_1012, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_1012, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_1013 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 1013)
        O2301172201202401690307101013 = {
            "NOMBRE": "Pago de Aportes para la ESAP personal docente",
            "CONCEPTO": "O2310102008 Aportes a la ESAP",
            "DISPONIBLE": df.loc[filtro_1013, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_1013, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_1013, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_1013, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_1013, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        # Total PARAFISCALES (1009-1013)
        filtro_parafiscales = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro.between(1009, 1013))
        PARAFISCALES = {
            "NOMBRE": "---",
            "CONCEPTO": "---",
            "DISPONIBLE": df.loc[filtro_parafiscales, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_parafiscales, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_parafiscales, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_parafiscales, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_parafiscales, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        # FOMAG (1014-1019)
        filtro_1014 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 1014)
        O2301172201202401690307101014 = {
            "NOMBRE": "Pago de Aportes para Salud del personal",
            "CONCEPTO": "O231010200201 Aportes a la seguridad social en salud publica",
            "DISPONIBLE": df.loc[filtro_1014, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_1014, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_1014, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_1014, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_1014, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_1015 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 1015)
        O2301172201202401690307101015 = {
            "NOMBRE": "Pago de Aportes para Pension del personal",
            "CONCEPTO": "O231010200101 Aportes a la seguridad social en pensiones publicas",
            "DISPONIBLE": df.loc[filtro_1015, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_1015, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_1015, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_1015, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_1015, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_1016 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 1016)
        O2301172201202401690307101016 = {
            "NOMBRE": "Pago de Aportes para Cesantias del personal",
            "CONCEPTO": "O231010200301 Aportes de cesantias a fondos publicos",
            "DISPONIBLE": df.loc[filtro_1016, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_1016, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_1016, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_1016, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_1016, "RECURSOS SIN EJECUTAR"].sum()
        }
        # Total FOMAG CSF (1014-1016)
        filtro_fomag_csf = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro.between(1014, 1016))
        FOMAG_CSF = {
            "NOMBRE": "---",
            "CONCEPTO": "---",
            "DISPONIBLE": df.loc[filtro_fomag_csf, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_fomag_csf, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_fomag_csf, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_fomag_csf, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_fomag_csf, "RECURSOS SIN EJECUTAR"].sum()
            }
        # FOMAG_Empleado_ssf (1017)
        filtro_1017 = (df["Codigo"].isin(["2-100-I002", "2-100-I001"])) & (ultimos_cuatro == 1017)
        O2301172201202401690307101017 = {
            "NOMBRE": "Pago de Personal Docente SSF primera inf",
            "CONCEPTO": "O231010100101 Sueldo basico",
            "DISPONIBLE": df.loc[filtro_1017, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_1017, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_1017, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_1017, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_1017, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        # FOMAG_SSF_Patron (1018)
        filtro_1018 = (df["Codigo"] == "2-100-I001") & (ultimos_cuatro == 1018)
        O2301172201202401690307101018 = {
            "NOMBRE": "Pago de Aportes para salud del personal",
            "CONCEPTO": "O231010200201 Aportes a la seguridad social en salud publica",
            "DISPONIBLE": df.loc[filtro_1018, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_1018, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_1018, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_1018, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_1018, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        # FOMAG_CSF (1019)
        filtro_1019 = (df["Codigo"] == "2-100-I001") & (ultimos_cuatro == 1019)
        O2301172201202401690307101019 = {
            "NOMBRE": "Pago de Aportes para Cesantias del perso",
            "CONCEPTO": "O231010200301 Aportes de cesantias a fondos publicos",
            "DISPONIBLE": df.loc[filtro_1019, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_1019, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_1019, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_1019, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_1019, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        # Total FOMAG SSF (1017-1019)
        filtro_fomag_ssf = (df["Codigo"].isin(["2-100-I002", "2-100-I001"])) & (ultimos_cuatro.between(1017, 1019))
        FOMAG_SSF = {
            "NOMBRE": "---",
            "CONCEPTO": "---",
            "DISPONIBLE": df.loc[filtro_fomag_ssf, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_fomag_ssf, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_fomag_ssf, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_fomag_ssf, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_fomag_ssf, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        # TOTAL_DOC_SGP (1001-1019)
        filtro_total = (df["Codigo"].isin(["2-100-I002", "2-100-I001", "1-204-I012"])) & (ultimos_cuatro.between(1001, 1019))
        TOTAL_DOC_SGP = {
            "NOMBRE": "---",
            "CONCEPTO": "---",
            "DISPONIBLE": df.loc[filtro_total, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_total, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_total, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_total, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_total, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        # --- 🔹 Crear tabla resumen ---
        resumen_principal = pd.DataFrame(
            [
                O2301172201202401690307101001, O2301172201202401690307101002, O2301172201202401690307101003,
                O2301172201202401690307101004, O2301172201202401690307101005, O2301172201202401690307101006,
                O2301172201202401690307101007, O2301172201202401690307101008,  
                SUELDOS,
                O2301172201202401690307101009, O2301172201202401690307101010, O2301172201202401690307101011,
                O2301172201202401690307101012, O2301172201202401690307101013,
                PARAFISCALES,
                O2301172201202401690307101014,O2301172201202401690307101015, O2301172201202401690307101016,
                FOMAG_CSF,
                O2301172201202401690307101017, O2301172201202401690307101018, O2301172201202401690307101019,
                FOMAG_SSF,
                TOTAL_DOC_SGP
            ],
            index=[
                "O2301172201202401690307101001", "O2301172201202401690307101002", "O2301172201202401690307101003",
                "O2301172201202401690307101004", "O2301172201202401690307101005", "O2301172201202401690307101006",
                "O2301172201202401690307101007", "O2301172201202401690307101008",  
                "SUELDOS",
                "O2301172201202401690307101009", "O2301172201202401690307101010", "O2301172201202401690307101011",
                "O2301172201202401690307101012", "O2301172201202401690307101013",
                "PARAFISCALES",
                "O2301172201202401690307101014","O2301172201202401690307101015", "O2301172201202401690307101016",
                "FOMAG_CSF",
                "O2301172201202401690307101017", "O2301172201202401690307101018", "O2301172201202401690307101019",
                "FOMAG_SSF",
                "TOTAL_DOC_SGP"
            ]
        )
        
        resumen = resumen_principal.copy()
        
        return df, resumen
         
    except Exception as e:
        st.error(f"❌ Error en procesar_datos_SGP_primera_infancia_detallada: {str(e)}")
        import traceback
        st.error(f"Detalle: {traceback.format_exc()}")
        return None, None

# =============================================================================
# FUNCIÓN DE PROCESAMIENTO SGP PRIMARIA BÁSICA Y MEDIA - DETALLADA (con NOMBRE y CONCEPTO)
# =============================================================================

def procesar_datos_SGP_primaria_basica_media_detallada(fuerza_actualizacion=False):
    """Función específica para procesar datos de SGP - Primaria Básica y Media (códigos 3033-3086) con detalle"""
    df = cargar_datos_originales(_fuerza_actualizacion=fuerza_actualizacion)
    
    if df is None:
        return None, None
    
    try:        
        ultimos_cuatro = pd.to_numeric(
            df["Codigo_O"].astype(str).str[-4:], 
            errors="coerce"
        )
        
        st.session_state.df_backup = df.copy()
        
        # --- 🔹 DEFINIR FILTROS INDIVIDUALES PARA PRIMARIA BÁSICA Y MEDIA ---
        
        # SUELDOS (3033-3041)
        filtro_3033 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3033)
        O2301172201202401690307103033 = {
            "NOMBRE": "Pago de Personal Docente",
            "CONCEPTO": "O231010100101 Sueldo basico",
            "DISPONIBLE": df.loc[filtro_3033, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3033, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3033, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3033, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3033, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3034 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3034)
        O2301172201202401690307103034 = {
            "NOMBRE": "Pago de Personal Directivo Docente",
            "CONCEPTO": "O231010100101 Sueldo basico",
            "DISPONIBLE": df.loc[filtro_3034, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3034, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3034, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3034, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3034, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3035 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3035)
        O2301172201202401690307103035 = {
            "NOMBRE": "Pago de Ascensos en escalafon del Personal",
            "CONCEPTO": "O231010100101 Sueldo basico",
            "DISPONIBLE": df.loc[filtro_3035, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3035, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3035, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3035, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3035, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3036 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3036)
        O2301172201202401690307103036 = {
            "NOMBRE": "Pago de horas extras del personal docent",
            "CONCEPTO": "O231010100102 Horas extras, dominicales, festivos y recargos",
            "DISPONIBLE": df.loc[filtro_3036, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3036, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3036, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3036, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3036, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3037 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3037)
        O2301172201202401690307103037 = {
            "NOMBRE": "Pago de Personal Docente- prima de servicio",
            "CONCEPTO": "O231010100106 Prima de servicio",
            "DISPONIBLE": df.loc[filtro_3037, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3037, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3037, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3037, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3037, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3038 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3038)
        O2301172201202401690307103038 = {
            "NOMBRE": "Pago de Personal Docente - prima de vacaciones",
            "CONCEPTO": "O23101010010802 Prima de vacaciones",
            "DISPONIBLE": df.loc[filtro_3038, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3038, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3038, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3038, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3038, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3039 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3039)
        O2301172201202401690307103039 = {
            "NOMBRE": "Pago de Personal Docente - prima de navidad",
            "CONCEPTO": "O23101010010801 Prima de navidad",
            "DISPONIBLE": df.loc[filtro_3039, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3039, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3039, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3039, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3039, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3040 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3040)
        O2301172201202401690307103040 = {
            "NOMBRE": "Pago de Personal Docente- subsidio de alimentación",
            "CONCEPTO": "O231010100104 Subsidio de alimentacion",
            "DISPONIBLE": df.loc[filtro_3040, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3040, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3040, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3040, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3040, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3041 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3041)
        O2301172201202401690307103041 = {
            "NOMBRE": "Pago Auxili de transporte personal docente",
            "CONCEPTO": "O231010100105 Auxilio de Transporte",
            "DISPONIBLE": df.loc[filtro_3041, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3041, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3041, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3041, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3041, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        # Total SUELDOS (3033-3041)
        filtro_sueldos = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro.between(3033, 3041))
        SUELDOS = {
            "NOMBRE": "---",
            "CONCEPTO": "---",
            "DISPONIBLE": df.loc[filtro_sueldos, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_sueldos, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_sueldos, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_sueldos, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_sueldos, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        # PARAFISCALES (3042-3051)
        filtro_3042 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3042)
        O2301172201202401690307103042 = {
            "NOMBRE": "Pago de Aportes para las Cajas de Compension",
            "CONCEPTO": "O231010200401 Compensar",
            "DISPONIBLE": df.loc[filtro_3042, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3042, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3042, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3042, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3042, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3043 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3043)
        O2301172201202401690307103043 = {
            "NOMBRE": "Pago de Aportes para las Cajas de Compensacion",
            "CONCEPTO": "O231010200401 Compensar",
            "DISPONIBLE": df.loc[filtro_3043, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3043, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3043, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3043, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3043, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3044 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3044)
        O2301172201202401690307103044 = {
            "NOMBRE": "Pago de Aportes para el ICBF personal docente",
            "CONCEPTO": "O2310102006 Aportes al ICBF",
            "DISPONIBLE": df.loc[filtro_3044, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3044, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3044, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3044, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3044, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3045 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3045)
        O2301172201202401690307103045 = {
            "NOMBRE": "Pago de Aportes para el ICBF del Personal",
            "CONCEPTO": "O2310102006 Aportes al ICBF",
            "DISPONIBLE": df.loc[filtro_3045, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3045, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3045, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3045, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3045, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3046 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3046)
        O2301172201202401690307103046 = {
            "NOMBRE": "Pago de Aportes para Institutos Tecnicos",
            "CONCEPTO": "O2310102009 Aportes a escuelas industriales e institutos tecnicos",
            "DISPONIBLE": df.loc[filtro_3046, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3046, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3046, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3046, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3046, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3047 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3047)
        O2301172201202401690307103047 = {
            "NOMBRE": "Pago de Aportes para Institutos Tecnicos",
            "CONCEPTO": "O2310102009 Aportes a escuelas industriales e institutos tecnicos",
            "DISPONIBLE": df.loc[filtro_3047, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3047, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3047, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3047, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3047, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3048 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3048)
        O2301172201202401690307103048 = {
            "NOMBRE": "Pago de Aportes para el SENA del Personal",
            "CONCEPTO": "O2310102007 Aportes al SENA",
            "DISPONIBLE": df.loc[filtro_3048, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3048, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3048, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3048, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3048, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3049 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3049)
        O2301172201202401690307103049 = {
            "NOMBRE": "Pago de Aportes para el SENA del Personal",
            "CONCEPTO": "O2310102007 Aportes al SENA",
            "DISPONIBLE": df.loc[filtro_3049, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3049, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3049, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3049, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3049, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3050 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3050)
        O2301172201202401690307103050 = {
            "NOMBRE": "Pago de Aportes para la ESAP personal docente",
            "CONCEPTO": "O2310102008 Aportes a la ESAP",
            "DISPONIBLE": df.loc[filtro_3050, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3050, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3050, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3050, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3050, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3051 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3051)
        O2301172201202401690307103051 = {
            "NOMBRE": "Pago de Aportes para la ESAP personal docente",
            "CONCEPTO": "O2310102008 Aportes a la ESAP",
            "DISPONIBLE": df.loc[filtro_3051, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3051, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3051, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3051, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3051, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        # Total PARAFISCALES (3042-3051)
        filtro_parafiscales = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro.between(3042, 3051))
        PARAFISCALES = {
            "NOMBRE": "---",
            "CONCEPTO": "---",
            "DISPONIBLE": df.loc[filtro_parafiscales, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_parafiscales, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_parafiscales, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_parafiscales, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_parafiscales, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        # FOMAG (3052-3063)
        filtro_3052 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3052)
        O2301172201202401690307103052 = {
            "NOMBRE": "Pago de Aportes para Salud del personal",
            "CONCEPTO": "O231010200201 Aportes a la seguridad social en salud publica",
            "DISPONIBLE": df.loc[filtro_3052, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3052, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3052, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3052, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3052, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3053 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3053)
        O2301172201202401690307103053 = {
            "NOMBRE": "Pago de Aportes para salud del personal",
            "CONCEPTO": "O231010200201 Aportes a la seguridad social en salud publica",
            "DISPONIBLE": df.loc[filtro_3053, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3053, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3053, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3053, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3053, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3054 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3054)
        O2301172201202401690307103054 = {
            "NOMBRE": "Pago de Aportes para Pension del personal",
            "CONCEPTO": "O231010200101 Aportes a la seguridad social en pensiones publicas",
            "DISPONIBLE": df.loc[filtro_3054, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3054, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3054, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3054, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3054, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3055 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3055)
        O2301172201202401690307103055 = {
            "NOMBRE": "Pago de Aportes para Pension del personal",
            "CONCEPTO": "O231010200101 Aportes a la seguridad social en pensiones publicas",
            "DISPONIBLE": df.loc[filtro_3055, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3055, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3055, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3055, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3055, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3056 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3056)
        O2301172201202401690307103056 = {
            "NOMBRE": "Pago de Aportes para Cesantias del personal",
            "CONCEPTO": "O231010200301 Aportes de cesantias a fondos publicos",
            "DISPONIBLE": df.loc[filtro_3056, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3056, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3056, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3056, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3056, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3057 = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro == 3057)
        O2301172201202401690307103057 = {
            "NOMBRE": "Pago de Aportes para Cesantias del personal",
            "CONCEPTO": "O231010200301 Aportes de cesantias a fondos publicos",
            "DISPONIBLE": df.loc[filtro_3057, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3057, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3057, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3057, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3057, "RECURSOS SIN EJECUTAR"].sum()
        }
        # Total FOMAG CSF (3052-3057)
        filtro_fomag_csf = (df["Codigo"] == "2-100-I002") & (ultimos_cuatro.between(3052, 3057))
        FOMAG_CSF = {
        "NOMBRE": "---",
        "CONCEPTO": "---",
        "DISPONIBLE": df.loc[filtro_fomag_csf, "DISPONIBLE"].sum(),
        "RP EMITIDOS": df.loc[filtro_fomag_csf, "RP EMITIDOS"].sum(),
        "GIROS ACUMULADOS": df.loc[filtro_fomag_csf, "GIROS ACUMULADOS"].sum(),
        "SALDO DE APROPIACION": df.loc[filtro_fomag_csf, "SALDO DE APROPIACION"].sum(),
        "RECURSOS SIN EJECUTAR": df.loc[filtro_fomag_csf, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        # FOMAG_Empleado (3058-3059) - Codigo 2-100-I002 y 2-100-I001
        filtro_3058 = (df["Codigo"].isin(["2-100-I002", "2-100-I001", "1-200-I038"])) & (ultimos_cuatro == 3058)
        O2301172201202401690307103058 = {
            "NOMBRE": "Pago de Personal Docente SSF",
            "CONCEPTO": "O231010100101 Sueldo basico",
            "DISPONIBLE": df.loc[filtro_3058, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3058, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3058, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3058, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3058, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3059 = (df["Codigo"].isin(["2-100-I002", "2-100-I001", "1-200-I038"])) & (ultimos_cuatro == 3059)
        O2301172201202401690307103059 = {
            "NOMBRE": "Pago de Personal Directivo  Docente SSF",
            "CONCEPTO": "O231010100101 Sueldo basico",
            "DISPONIBLE": df.loc[filtro_3059, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3059, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3059, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3059, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3059, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        # FOMAG_SSF_Patron (3060-3063) - Codigo 2-100-I001
        filtro_3060 = (df["Codigo"].isin(["2-100-I002", "2-100-I001", "1-200-I038"])) & (ultimos_cuatro == 3060)
        O2301172201202401690307103060 = {
            "NOMBRE": "Pago de Aportes para salud del personal",
            "CONCEPTO": "O231010200201 Aportes a la seguridad social en salud publica",
            "DISPONIBLE": df.loc[filtro_3060, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3060, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3060, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3060, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3060, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3061 = (df["Codigo"].isin(["2-100-I002", "2-100-I001", "1-200-I038"])) & (ultimos_cuatro == 3061)
        O2301172201202401690307103061 = {
            "NOMBRE": "Pago de Aportes para salud del personal",
            "CONCEPTO": "O231010200201 Aportes a la seguridad social en salud publica",
            "DISPONIBLE": df.loc[filtro_3061, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3061, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3061, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3061, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3061, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3062 = (df["Codigo"].isin(["2-100-I002", "2-100-I001", "1-200-I038"])) & (ultimos_cuatro == 3062)
        O2301172201202401690307103062 = {
            "NOMBRE": "Pago de Aportes para Cesantias del personal",
            "CONCEPTO": "O231010200301 Aportes de cesantias a fondos publicos",
            "DISPONIBLE": df.loc[filtro_3062, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3062, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3062, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3062, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3062, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        filtro_3063 = (df["Codigo"] == "2-100-I001") & (ultimos_cuatro == 3063)
        O2301172201202401690307103063 = {
            "NOMBRE": "Pago de Aportes para Cesantias del personal",
            "CONCEPTO": "O231010200301 Aportes de cesantias a fondos publicos",
            "DISPONIBLE": df.loc[filtro_3063, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_3063, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_3063, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_3063, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_3063, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        # Total FOMAG SSF
        filtro_fomag_ssf = (df["Codigo"].isin(["2-100-I002", "2-100-I001", "1-200-I038"])) & (ultimos_cuatro.between(3058, 3063))
        FOMAG_SSF = {
            "NOMBRE": "---",
            "CONCEPTO": "---",
            "DISPONIBLE": df.loc[filtro_fomag_ssf, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_fomag_ssf, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_fomag_ssf, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_fomag_ssf, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_fomag_ssf, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        
        # TOTAL_SGP_P8033 (3033-3063)
        filtro_total = (df["Codigo"].isin(["2-100-I002", "2-100-I001", "2-100-I026", "1-204-I012", "1-200-I038"])) & (ultimos_cuatro.between(3033, 3063))
        TOTAL_SGP_P8033 = {
            "NOMBRE": "---",
            "CONCEPTO": "---",
            "DISPONIBLE": df.loc[filtro_total, "DISPONIBLE"].sum(),
            "RP EMITIDOS": df.loc[filtro_total, "RP EMITIDOS"].sum(),
            "GIROS ACUMULADOS": df.loc[filtro_total, "GIROS ACUMULADOS"].sum(),
            "SALDO DE APROPIACION": df.loc[filtro_total, "SALDO DE APROPIACION"].sum(),
            "RECURSOS SIN EJECUTAR": df.loc[filtro_total, "RECURSOS SIN EJECUTAR"].sum()
        }
        
        # --- 🔹 Crear tabla resumen ---
        resumen_principal = pd.DataFrame(
            [
                O2301172201202401690307103033, O2301172201202401690307103034, O2301172201202401690307103035,
                O2301172201202401690307103036, O2301172201202401690307103037, O2301172201202401690307103038,
                O2301172201202401690307103039, O2301172201202401690307103040, O2301172201202401690307103041,  
                SUELDOS,
                O2301172201202401690307103042, O2301172201202401690307103043, O2301172201202401690307103044,
                O2301172201202401690307103045, O2301172201202401690307103046, O2301172201202401690307103047,
                O2301172201202401690307103048, O2301172201202401690307103049, O2301172201202401690307103050,
                O2301172201202401690307103051,
                PARAFISCALES,
                O2301172201202401690307103052, O2301172201202401690307103053, O2301172201202401690307103054,
                O2301172201202401690307103055, O2301172201202401690307103056, O2301172201202401690307103057,
                FOMAG_CSF,
                O2301172201202401690307103058, O2301172201202401690307103059, O2301172201202401690307103060,
                O2301172201202401690307103061, O2301172201202401690307103062, O2301172201202401690307103063,
                FOMAG_SSF,
                TOTAL_SGP_P8033
            ],
            index=[
                "O2301172201202401690307103033", "O2301172201202401690307103034", "O2301172201202401690307103035",
                "O2301172201202401690307103036", "O2301172201202401690307103037", "O2301172201202401690307103038",
                "O2301172201202401690307103039", "O2301172201202401690307103040", "O2301172201202401690307103041",  
                "SUELDOS",
                "O2301172201202401690307103042", "O2301172201202401690307103043", "O2301172201202401690307103044",
                "O2301172201202401690307103045", "O2301172201202401690307103046", "O2301172201202401690307103047",
                "O2301172201202401690307103048", "O2301172201202401690307103049", "O2301172201202401690307103050",
                "O2301172201202401690307103051",
                "PARAFISCALES",
                "O2301172201202401690307103052", "O2301172201202401690307103053", "O2301172201202401690307103054",
                "O2301172201202401690307103055", "O2301172201202401690307103056", "O2301172201202401690307103057",
                "FOMAG_CSF",
                "O2301172201202401690307103058", "O2301172201202401690307103059", "O2301172201202401690307103060",
                "O2301172201202401690307103061", "O2301172201202401690307103062", "O2301172201202401690307103063",
                "FOMAG_SSF",
                "TOTAL_SGP_P8033"
            ]
        )
        
        resumen = resumen_principal.copy()
        
        return df, resumen
         
    except Exception as e:
        st.error(f"❌ Error en procesar_datos_SGP_primaria_basica_media_detallada: {str(e)}")
        import traceback
        st.error(f"Detalle: {traceback.format_exc()}")
        return None, None


# =============================================================================
# FUNCIONES DE VISUALIZACIÓN
# =============================================================================

def mostrar_tabla_sgp(resumen):
    if resumen is None or resumen.empty:
        st.warning("No hay datos para mostrar")
        return

    resumen_formateado = resumen.copy()
    for col in resumen_formateado.columns:
        resumen_formateado[col] = resumen_formateado[col].apply(
            lambda x: f"${x:,.0f}".replace(",", ".")
            if pd.notnull(x) and isinstance(x, (int, float)) else "$0"
        )

    st.markdown(
        "<div class='titulo-tabla'>📊 TABLA RESUMEN EJECUCIÓN PRESUPUESTAL</div>",
        unsafe_allow_html=True
    )

    html_tabla = """
<div class="tabla-container">
<table class="tabla-personalizada">
<thead>
<tr>
<th>BOLSILLOS</th>
<th>INICIAL</th>
<th>DISPONIBLE</th>
<th>RP EMITIDOS</th>
<th>GIROS ACUMULADOS</th>
<th>SALDO DE APROPIACION</th>
<th>RECURSOS SIN EJECUTAR</th>
</tr>
</thead>
<tbody>
"""
    filas_totales = {
        "TOTAL_DOC_SGP": "fila-total",
        "TOTAL_SGP_P8033": "fila-total-final",
        "TOTAL_RP_P8033": "fila-total-final",
        "TOTAL_GENERAL": "fila-total-general"
    }

    for idx, row in resumen_formateado.iterrows():
        clase_fila = filas_totales.get(idx, "")
        html_tabla += f"""
<tr class="{clase_fila}">
<td class="encabezado-fila">{idx}</td>
<td class="numero">{row['INICIAL']}</td>
<td class="numero">{row['DISPONIBLE']}</td>
<td class="numero">{row['RP EMITIDOS']}</td>
<td class="numero">{row['GIROS ACUMULADOS']}</td>
<td class="numero">{row['SALDO DE APROPIACION']}</td>
<td class="numero">{row['RECURSOS SIN EJECUTAR']}</td>
</tr>
"""

    html_tabla += """
</tbody>
</table>
</div>
"""

    st.markdown(html_tabla, unsafe_allow_html=True)


# =============================================================================
# FUNCIÓN DE VISUALIZACIÓN DOCENTES RP PRINCIPAL 
# =============================================================================

def mostrar_tabla_RP_principal(resumen):
    if resumen is None or resumen.empty:
        st.warning("No hay datos para mostrar")
        return

    resumen_formateado = resumen.copy()
    for col in resumen_formateado.columns:
        resumen_formateado[col] = resumen_formateado[col].apply(
            lambda x: f"${x:,.0f}".replace(",", ".")
            if pd.notnull(x) and isinstance(x, (int, float)) else "$0"
        )

    st.markdown(
        "<div class='titulo-tabla'>📊 TABLA RESUMEN EJECUCIÓN PRESUPUESTAL DOCENTES RP</div>",
        unsafe_allow_html=True
    )

    html_tabla = """
<div class="tabla-container">
<table class="tabla-personalizada">
<thead>
<tr>
<th>OBJETO</th>
<th>DISPONIBLE</th>
<th>RP EMITIDOS</th>
<th>GIROS ACUMULADOS</th>
<th>SALDO DE APROPIACION</th>
<th>RECURSOS SIN EJECUTAR</th>
</tr>
</thead>
<tbody>
"""
    filas_totales = {
        "SUELDOS": "fila-total", 
        "PARAFISCALES": "fila-total-final",
        "FOMAG": "fila-total-general","TOTAL_DOC_RP":"fila-total" 
    }

    for idx, row in resumen_formateado.iterrows():
        clase_fila = filas_totales.get(idx, "")
        html_tabla += f"""
<tr class="{clase_fila}">
<td class="encabezado-fila">{idx}</td>
<td class="numero">{row['DISPONIBLE']}</td>
<td class="numero">{row['RP EMITIDOS']}</td>
<td class="numero">{row['GIROS ACUMULADOS']}</td>
<td class="numero">{row['SALDO DE APROPIACION']}</td>
<td class="numero">{row['RECURSOS SIN EJECUTAR']}</td>
</tr>
"""

    html_tabla += """
</tbody>
</table>
</div>
"""

    st.markdown(html_tabla, unsafe_allow_html=True)



# =============================================================================
# FUNCIÓN DE VISUALIZACIÓN DOCENTES RP PRIMERA INFANCIA - CORREGIDA
# =============================================================================

def mostrar_tabla_RP_primera_infancia(resumen):
    if resumen is None or resumen.empty:
        st.warning("No hay datos para mostrar")
        return

    # Crear una copia para formatear
    resumen_formateado = resumen.copy()
    
    # Formatear SOLO las columnas numéricas (no las de texto)
    columnas_numericas = ['DISPONIBLE', 'RP EMITIDOS', 'GIROS ACUMULADOS', 
                         'SALDO DE APROPIACION', 'RECURSOS SIN EJECUTAR']
    
    for col in columnas_numericas:
        if col in resumen_formateado.columns:
            resumen_formateado[col] = resumen_formateado[col].apply(
                lambda x: f"${x:,.0f}".replace(",", ".")
                if pd.notnull(x) and isinstance(x, (int, float)) else "$0"
            )

    st.markdown(
        "<div class='titulo-tabla'>📊 EJECUCIÓN PRESUPUESTAL DOCENTES PRIMERA INFANCIA RP</div>",
        unsafe_allow_html=True
    )

    # HTML CON CLASES ESPECÍFICAS PARA COLUMNAS DE TEXTO
    html_tabla = """
<div class="tabla-container">
<table class="tabla-personalizada">
<thead>
<tr>
<th class="encabezado-fila">CÓDIGO</th>
<th class="encabezado-texto">NOMBRE</th>
<th class="encabezado-texto">CONCEPTO</th>
<th>DISPONIBLE</th>
<th>RP EMITIDOS</th>
<th>GIROS ACUMULADOS</th>
<th>SALDO DE APROPIACION</th>
<th>RECURSOS SIN EJECUTAR</th>
</tr>
</thead>
<tbody>
"""
    filas_totales = {
        "SUELDOS": "fila-total", 
        "PARAFISCALES": "fila-total-final",
        "FOMAG": "fila-total-general", "TOTAL_PRIMERA_INFANCIA": "fila-total"
    }
    for idx, row in resumen_formateado.iterrows():
        clase_fila = filas_totales.get(idx, "")
        html_tabla += f"""
<tr>
<tr class="{clase_fila}">
<td class="encabezado-fila">{idx}</td>
<td class="texto-reducido">{row.get('NOMBRE', '-')}</td>
<td class="texto-reducido">{row.get('CONCEPTO', '-')}</td>
<td class="numero">{row.get('DISPONIBLE', '$0')}</td>
<td class="numero">{row.get('RP EMITIDOS', '$0')}</td>
<td class="numero">{row.get('GIROS ACUMULADOS', '$0')}</td>
<td class="numero">{row.get('SALDO DE APROPIACION', '$0')}</td>
<td class="numero">{row.get('RECURSOS SIN EJECUTAR', '$0')}</td>
</tr>
"""

    html_tabla += """
</tbody>
</table>
</div>
"""

    st.markdown(html_tabla, unsafe_allow_html=True)


# =============================================================================
# FUNCIÓN DE VISUALIZACIÓN DOCENTES RP ORIENTADORES
# =============================================================================

def mostrar_tabla_RP_orientadores(resumen):
    if resumen is None or resumen.empty:
        st.warning("No hay datos para mostrar")
        return

    # Crear una copia para formatear
    resumen_formateado = resumen.copy()
    
    # Formatear SOLO las columnas numéricas (no las de texto)
    columnas_numericas = ['DISPONIBLE', 'RP EMITIDOS', 'GIROS ACUMULADOS', 
                         'SALDO DE APROPIACION', 'RECURSOS SIN EJECUTAR']
    
    for col in columnas_numericas:
        if col in resumen_formateado.columns:
            resumen_formateado[col] = resumen_formateado[col].apply(
                lambda x: f"${x:,.0f}".replace(",", ".")
                if pd.notnull(x) and isinstance(x, (int, float)) else "$0"
            )

    st.markdown(
        "<div class='titulo-tabla'>📊 EJECUCIÓN PRESUPUESTAL DOCENTES ORIENTADORES RP</div>",
        unsafe_allow_html=True
    )

    # HTML CON CLASES ESPECÍFICAS PARA COLUMNAS DE TEXTO
    html_tabla = """
<div class="tabla-container">
<table class="tabla-personalizada">
<thead>
<tr>
<th class="encabezado-fila">CÓDIGO</th>
<th class="encabezado-texto">NOMBRE</th>
<th class="encabezado-texto">CONCEPTO</th>
<th>DISPONIBLE</th>
<th>RP EMITIDOS</th>
<th>GIROS ACUMULADOS</th>
<th>SALDO DE APROPIACION</th>
<th>RECURSOS SIN EJECUTAR</th>
</tr>
</thead>
<tbody>
"""
    filas_totales = {
        "SUELDOS_ORIENTADORES": "fila-total", 
        "PARAFISCALES_ORIENTADORES": "fila-total-final",
        "FOMAG_ORIENTADORES": "fila-total-general", "TOTAL_DOC_ORIENTADORES": "fila-total"
    }
    for idx, row in resumen_formateado.iterrows():
        clase_fila = filas_totales.get(idx, "")
        html_tabla += f"""
<tr>
<tr class="{clase_fila}">
<td class="encabezado-fila">{idx}</td>
<td class="texto-reducido">{row.get('NOMBRE', '-')}</td>
<td class="texto-reducido">{row.get('CONCEPTO', '-')}</td>
<td class="numero">{row.get('DISPONIBLE', '$0')}</td>
<td class="numero">{row.get('RP EMITIDOS', '$0')}</td>
<td class="numero">{row.get('GIROS ACUMULADOS', '$0')}</td>
<td class="numero">{row.get('SALDO DE APROPIACION', '$0')}</td>
<td class="numero">{row.get('RECURSOS SIN EJECUTAR', '$0')}</td>
</tr>
"""

    html_tabla += """
</tbody>
</table>
</div>
"""

    st.markdown(html_tabla, unsafe_allow_html=True)

# =============================================================================
# FUNCIÓN DE VISUALIZACIÓN DOCENTES RP PRIMARIA BASICA MEDIA
# =============================================================================

def mostrar_tabla_RP_PBM(resumen):
    if resumen is None or resumen.empty:
        st.warning("No hay datos para mostrar")
        return

    # Crear una copia para formatear
    resumen_formateado = resumen.copy()
    
    # Formatear SOLO las columnas numéricas (no las de texto)
    columnas_numericas = ['DISPONIBLE', 'RP EMITIDOS', 'GIROS ACUMULADOS', 
                         'SALDO DE APROPIACION', 'RECURSOS SIN EJECUTAR']
    
    for col in columnas_numericas:
        if col in resumen_formateado.columns:
            resumen_formateado[col] = resumen_formateado[col].apply(
                lambda x: f"${x:,.0f}".replace(",", ".")
                if pd.notnull(x) and isinstance(x, (int, float)) else "$0"
            )

    st.markdown(
        "<div class='titulo-tabla'>📊 EJECUCIÓN PRESUPUESTAL DOCENTES PRIMARIA BASICA MEDIA RP</div>",
        unsafe_allow_html=True
    )

    # HTML CON CLASES ESPECÍFICAS PARA COLUMNAS DE TEXTO
    html_tabla = """
<div class="tabla-container">
<table class="tabla-personalizada">
<thead>
<tr>
<th class="encabezado-fila">CÓDIGO</th>
<th class="encabezado-texto">NOMBRE</th>
<th class="encabezado-texto">CONCEPTO</th>
<th>DISPONIBLE</th>
<th>RP EMITIDOS</th>
<th>GIROS ACUMULADOS</th>
<th>SALDO DE APROPIACION</th>
<th>RECURSOS SIN EJECUTAR</th>
</tr>
</thead>
<tbody>
"""
    filas_totales = {
        "SUELDOS_PBM": "fila-total", 
        "PARAFISCALES_PBM": "fila-total-final",
        "FOMAG_PBM": "fila-total-general", "TOTAL_DOC_PBM": "fila-total"
    }
    for idx, row in resumen_formateado.iterrows():
        clase_fila = filas_totales.get(idx, "")
        html_tabla += f"""
<tr>
<tr class="{clase_fila}">
<td class="encabezado-fila">{idx}</td>
<td class="texto-reducido">{row.get('NOMBRE', '-')}</td>
<td class="texto-reducido">{row.get('CONCEPTO', '-')}</td>
<td class="numero">{row.get('DISPONIBLE', '$0')}</td>
<td class="numero">{row.get('RP EMITIDOS', '$0')}</td>
<td class="numero">{row.get('GIROS ACUMULADOS', '$0')}</td>
<td class="numero">{row.get('SALDO DE APROPIACION', '$0')}</td>
<td class="numero">{row.get('RECURSOS SIN EJECUTAR', '$0')}</td>
</tr>
"""

    html_tabla += """
</tbody>
</table>
</div>
"""

    st.markdown(html_tabla, unsafe_allow_html=True)

# =============================================================================
# FUNCIÓN DE VISUALIZACIÓN DOCENTES SGP PRINCIPAL 
# =============================================================================

def mostrar_tabla_SGPP_principal(resumen):
    if resumen is None or resumen.empty:
        st.warning("No hay datos para mostrar")
        return

    resumen_formateado = resumen.copy()
    for col in resumen_formateado.columns:
        resumen_formateado[col] = resumen_formateado[col].apply(
            lambda x: f"${x:,.0f}".replace(",", ".")
            if pd.notnull(x) and isinstance(x, (int, float)) else "$0"
        )

    st.markdown(
        "<div class='titulo-tabla'>📊 TABLA RESUMEN EJECUCIÓN PRESUPUESTAL DOCENTES SGP</div>",
        unsafe_allow_html=True
    )

    html_tabla = """
<div class="tabla-container">
<table class="tabla-personalizada">
<thead>
<tr>
<th>OBJETO</th>
<th>DISPONIBLE</th>
<th>RP EMITIDOS</th>
<th>GIROS ACUMULADOS</th>
<th>SALDO DE APROPIACION</th>
<th>RECURSOS SIN EJECUTAR</th>
</tr>
</thead>
<tbody>
"""
    filas_totales = {
        "SUELDOS": "fila-total", 
        "PARAFISCALES": "fila-total-final",
        "FOMAG_ssf": "fila-total-general","FOMAG_csf":"fila-total", "TOTAL_DOC_SGP":"fila-total"     
    }

    for idx, row in resumen_formateado.iterrows():
        clase_fila = filas_totales.get(idx, "")
        html_tabla += f"""
<tr class="{clase_fila}">
<td class="encabezado-fila">{idx}</td>
<td class="numero">{row['DISPONIBLE']}</td>
<td class="numero">{row['RP EMITIDOS']}</td>
<td class="numero">{row['GIROS ACUMULADOS']}</td>
<td class="numero">{row['SALDO DE APROPIACION']}</td>
<td class="numero">{row['RECURSOS SIN EJECUTAR']}</td>
</tr>
"""

    html_tabla += """
</tbody>
</table>
</div>
"""

    st.markdown(html_tabla, unsafe_allow_html=True)

# =============================================================================
# FUNCIÓN DE VISUALIZACIÓN SGP PRIMERA INFANCIA - DETALLADA
# =============================================================================

def mostrar_tabla_SGP_primera_infancia_detallada(resumen):
    if resumen is None or resumen.empty:
        st.warning("No hay datos para mostrar")
        return

    # Crear una copia para formatear
    resumen_formateado = resumen.copy()
    
    # Formatear SOLO las columnas numéricas
    columnas_numericas = ['DISPONIBLE', 'RP EMITIDOS', 'GIROS ACUMULADOS', 
                         'SALDO DE APROPIACION', 'RECURSOS SIN EJECUTAR']
    
    for col in columnas_numericas:
        if col in resumen_formateado.columns:
            resumen_formateado[col] = resumen_formateado[col].apply(
                lambda x: f"${x:,.0f}".replace(",", ".")
                if pd.notnull(x) and isinstance(x, (int, float)) else "$0"
            )

    st.markdown(
        "<div class='titulo-tabla'>📊 EJECUCIÓN PRESUPUESTAL DOCENTES PRIMERA INFANCIA SGP</div>",
        unsafe_allow_html=True
    )

    html_tabla = """
<div class="tabla-container">
<table class="tabla-personalizada">
<thead>
<tr>
<th class="encabezado-fila">CÓDIGO</th>
<th class="encabezado-texto">NOMBRE</th>
<th class="encabezado-texto">CONCEPTO</th>
<th>DISPONIBLE</th>
<th>RP EMITIDOS</th>
<th>GIROS ACUMULADOS</th>
<th>SALDO DE APROPIACION</th>
<th>RECURSOS SIN EJECUTAR</th>
</tr>
</thead>
<tbody>
"""
    filas_totales = {
        "SUELDOS": "fila-total", 
        "PARAFISCALES": "fila-total-final",
        "FOMAG_CSF": "fila-total-general",
        "FOMAG_SSF": "fila-total-general",
        "TOTAL_DOC_SGP": "fila-total"
    }
    
    for idx, row in resumen_formateado.iterrows():
        clase_fila = filas_totales.get(idx, "")
        html_tabla += f"""
<tr class="{clase_fila}">
<td class="encabezado-fila">{idx}</td>
<td class="texto-reducido">{row.get('NOMBRE', '-')}</td>
<td class="texto-reducido">{row.get('CONCEPTO', '-')}</td>
<td class="numero">{row.get('DISPONIBLE', '$0')}</td>
<td class="numero">{row.get('RP EMITIDOS', '$0')}</td>
<td class="numero">{row.get('GIROS ACUMULADOS', '$0')}</td>
<td class="numero">{row.get('SALDO DE APROPIACION', '$0')}</td>
<td class="numero">{row.get('RECURSOS SIN EJECUTAR', '$0')}</td>
</tr>
"""

    html_tabla += """
</tbody>
</table>
</div>
"""

    st.markdown(html_tabla, unsafe_allow_html=True)


# =============================================================================
# FUNCIÓN DE VISUALIZACIÓN SGP PRIMARIA BÁSICA Y MEDIA - DETALLADA
# =============================================================================

def mostrar_tabla_SGP_primaria_basica_media_detallada(resumen):
    if resumen is None or resumen.empty:
        st.warning("No hay datos para mostrar")
        return

    # Crear una copia para formatear
    resumen_formateado = resumen.copy()
    
    # Formatear SOLO las columnas numéricas
    columnas_numericas = ['DISPONIBLE', 'RP EMITIDOS', 'GIROS ACUMULADOS', 
                         'SALDO DE APROPIACION', 'RECURSOS SIN EJECUTAR']
    
    for col in columnas_numericas:
        if col in resumen_formateado.columns:
            resumen_formateado[col] = resumen_formateado[col].apply(
                lambda x: f"${x:,.0f}".replace(",", ".")
                if pd.notnull(x) and isinstance(x, (int, float)) else "$0"
            )

    st.markdown(
        "<div class='titulo-tabla'>📊 EJECUCIÓN PRESUPUESTAL DOCENTES PRIMARIA BÁSICA MEDIA SGP</div>",
        unsafe_allow_html=True
    )

    html_tabla = """
<div class="tabla-container">
<table class="tabla-personalizada">
<thead>
<tr>
<th class="encabezado-fila">CÓDIGO</th>
<th class="encabezado-texto">NOMBRE</th>
<th class="encabezado-texto">CONCEPTO</th>
<th>DISPONIBLE</th>
<th>RP EMITIDOS</th>
<th>GIROS ACUMULADOS</th>
<th>SALDO DE APROPIACION</th>
<th>RECURSOS SIN EJECUTAR</th>
</tr>
</thead>
<tbody>
"""
    filas_totales = {
        "SUELDOS": "fila-total", 
        "PARAFISCALES": "fila-total-final",
        "FOMAG_CSF": "fila-total-general",
        "FOMAG_SSF": "fila-total-general", 
        "TOTAL_SGP_P8033": "fila-total"
    }
    
    for idx, row in resumen_formateado.iterrows():
        clase_fila = filas_totales.get(idx, "")
        html_tabla += f"""
<tr class="{clase_fila}">
<td class="encabezado-fila">{idx}</td>
<td class="texto-reducido">{row.get('NOMBRE', '-')}</td>
<td class="texto-reducido">{row.get('CONCEPTO', '-')}</td>
<td class="numero">{row.get('DISPONIBLE', '$0')}</td>
<td class="numero">{row.get('RP EMITIDOS', '$0')}</td>
<td class="numero">{row.get('GIROS ACUMULADOS', '$0')}</td>
<td class="numero">{row.get('SALDO DE APROPIACION', '$0')}</td>
<td class="numero">{row.get('RECURSOS SIN EJECUTAR', '$0')}</td>
</tr>
"""

    html_tabla += """
</tbody>
</table>
</div>
"""

    st.markdown(html_tabla, unsafe_allow_html=True)



# =============================================================================
# PANTALLAS
# =============================================================================
def mostrar_pantalla_inicial():
    st.markdown("<div class='pantalla-inicial'>", unsafe_allow_html=True)
    
    # Logos
    st.markdown("<div class='contenedor-logos'>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 2, 1])
    with col1:
        st.image("logo_bogota.png", width=250, use_container_width=True)
    with col2:
        st.markdown("<div class='header'><h1>SECRETARÍA DE EDUCACIÓN DE BOGOTÁ</h1><h2>CONTROL PRESUPUESTAL NÓMINA</h2></div>", unsafe_allow_html=True)
    with col3:
        st.image("logo_alcaldía_mayor.png", width=250, use_container_width=True)
    st.markdown("</div>", unsafe_allow_html=True)
    # 🔹 NUEVO: Mostrar fecha de actualización en la pantalla inicial
    mostrar_fecha_actualizacion()
    
    # Botones
    st.markdown("<div class='contenedor-botones'>", unsafe_allow_html=True)
    
    col_principal = st.columns([1])
    with col_principal[0]:
        if st.button("🏠 TABLERO PRINCIPAL", key="principal", use_container_width=True):
            st.session_state.pagina_actual = "POR_FUENTE"
            st.rerun()
    
    st.markdown("<div style='margin: 20px 0;'></div>", unsafe_allow_html=True)
    
    col_secundarios = st.columns(2)
    with col_secundarios[0]:
        if st.button("💰 RECURSOS PROPIOS", key="recursos_propios", use_container_width=True):
            st.session_state.pagina_actual = "RECURSOS_PROPIOS"
            st.rerun()
    with col_secundarios[1]:
        if st.button("📊 SGP", key="sgp", use_container_width=True):
            st.session_state.pagina_actual = "SGP"
            st.rerun()
    
    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

def mostrar_pantalla_por_fuente():
    # Botón de volver
    if st.button("← Volver al Inicio", key="volver_fuente"):
        st.session_state.pagina_actual = "INICIO"
        st.rerun()
    
    # HEADER MEJORADO
    st.markdown("<div class='contenedor-logos'>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col1:
        st.image("logo_bogota.png", width=70, use_container_width=True)
    
    with col2:
        st.markdown("""
        <div class='header'>
            <h2>🏢 TABLERO PRESUPUESTAL</h2>
            <p style='margin:0; font-size:14px;'>Actualizado desde Google Sheets</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.image("logo_alcaldía_mayor.png", width=70, use_container_width=True)
    st.markdown("</div>", unsafe_allow_html=True)
    
    # Mostrar fecha de actualización
    mostrar_fecha_actualizacion()
    
    # BOTÓN DE ACTUALIZACIÓN PROMINENTE
    col_boton, col_espacio = st.columns([1, 5])
    with col_boton:
        if st.button("🔄 **ACTUALIZAR**", 
                    type="primary",
                    use_container_width=True,
                    help="Forzar actualización inmediata desde Google Sheets"):
            # Forzar recarga
            st.cache_data.clear()  # Limpiar cache
            st.session_state.force_update = True
            st.rerun()
    
    # Cargar datos (con o sin fuerza)
    fuerza = st.session_state.get('force_update', False)
    
    with st.spinner("📊 Procesando datos presupuestales..."):
        df, resumen = procesar_datos_sgp(fuerza_actualizacion=fuerza)
        
        # Resetear flag
        if 'force_update' in st.session_state:
            st.session_state.force_update = False
    
    # Mostrar resultados
    if resumen is not None:
        mostrar_tabla_sgp(resumen)
        
        # =========================================================================
        # 🔹 BOTÓN DE DESCARGA EXCEL - TABLERO PRINCIPAL
        # =========================================================================
        st.divider()
        st.subheader("📥 Exportar a Excel")
        
        # Descarga del Tablero Principal
        excel_tablero = exportar_a_excel_formateado(
            [resumen], 
            ["TABLERO_PRINCIPAL"], 
            ["SGP"]
        )
        
        st.download_button(
            label="📥 Descargar Tablero Principal",
            data=excel_tablero,
            file_name=f"TABLERO_PRINCIPAL_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
            help="Descarga la tabla del tablero principal con todos los formatos y colores"
        )
        
        # INFO DE ACTUALIZACIÓN
        st.divider()
        col_info1, col_info2 = st.columns(2)
        with col_info1:
            fecha = st.session_state.get('fecha_actualizacion', datetime.now().strftime('%d/%m/%Y %H:%M'))
            st.caption(f"📅 {fecha}")
        with col_info2:
            st.caption(f"📊 {len(df)} registros cargados" if df is not None else "📊 Datos no disponibles")
    else:
        st.error("No se pudieron cargar los datos. Intenta actualizar.")


# =============================================================================
# PANTALLA DOCENTES RP
# =============================================================================


def mostrar_pantalla_recursos_propios():
    if st.button("← Volver al Inicio", key="volver_recursos"):
        st.session_state.pagina_actual = "INICIO"
        st.rerun()
    
    st.markdown("<div class='contenedor-logos'>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 2, 1])
    with col1:
        st.image("logo_bogota.png", width=150, use_container_width=True)
    with col2:
        st.markdown("<div class='header'><h2>DOCENTES RECURSOS PROPIOS</h2></div>", unsafe_allow_html=True)
    with col3:
        st.image("logo_alcaldía_mayor.png", width=150, use_container_width=True)
    st.markdown("</div>", unsafe_allow_html=True)

     # 🔹 NUEVO: Mostrar fecha de actualización
    mostrar_fecha_actualizacion()
    
    # =========================================================================
    # BOTÓN DE ACTUALIZACIÓN GLOBAL (funciona para todas las secciones)
    # =========================================================================
    col_actualizar, _ = st.columns([1, 3])
    with col_actualizar:
        if st.button("🔄 Actualizar TODOS los Datos", key="actualizar_todos", use_container_width=True):
            st.cache_data.clear()
            st.session_state.force_update_all = True
            st.rerun()
    
    # Usar la misma bandera para todas las cargas
    fuerza = st.session_state.get('force_update_all', False)
    
    # =========================================================================
    # 1. TOTAL
    # =========================================================================
    st.subheader("🌐 TOTAL")
    
    # Botón específico para esta sección
    col_total, _ = st.columns([1, 3])
    with col_total:
        if st.button("🔄 Actualizar TOTAL", key="actualizar_total", use_container_width=True):
            st.cache_data.clear()
            st.session_state.force_update_total = True
            st.rerun()
    
    fuerza_total = st.session_state.get('force_update_total', fuerza)
    
    with st.spinner("📊 Procesando datos TOTAL..."):
        df, resumen = procesar_datos_RP_principal(fuerza_actualizacion=fuerza_total)
        
        if 'force_update_total' in st.session_state:
            st.session_state.force_update_total = False
    
    if df is not None:
        if resumen is not None:
            mostrar_tabla_RP_principal(resumen)
        else:
            st.error("❌ No se pudieron procesar los datos TOTAL")
    else:
        st.error("❌ No se pudo cargar el archivo de datos")
    
    st.divider()
    
    # =========================================================================
    # 2. PRIMERA INFANCIA
    # =========================================================================
    st.subheader("👶🏻 Primera Infancia")
    
    # Botón específico para esta sección
    col_pi, _ = st.columns([1, 3])
    with col_pi:
        if st.button("🔄 Actualizar Primera Infancia", key="actualizar_pi", use_container_width=True):
            st.cache_data.clear()
            st.session_state.force_update_pi = True
            st.rerun()
    
    fuerza_pi = st.session_state.get('force_update_pi', fuerza)
    
    with st.spinner("📊 Procesando datos Primera Infancia..."):
        df, resumen = procesar_datos_RP_primera_infancia(fuerza_actualizacion=fuerza_pi)
        
        if 'force_update_pi' in st.session_state:
            st.session_state.force_update_pi = False
    
    if df is not None:
        if resumen is not None:
            mostrar_tabla_RP_primera_infancia(resumen)
        else:
            st.error("❌ No se pudieron procesar los datos Primera Infancia")
    else:
        st.error("❌ No se pudo cargar el archivo de datos")
    
    st.divider()
    
    # =========================================================================
    # 3. ORIENTADORES
    # =========================================================================
    st.subheader("🤝🏻 Orientadores")
    
    # Botón específico para esta sección
    col_or, _ = st.columns([1, 3])
    with col_or:
        if st.button("🔄 Actualizar Orientadores", key="actualizar_or", use_container_width=True):
            st.cache_data.clear()
            st.session_state.force_update_or = True
            st.rerun()
    
    fuerza_or = st.session_state.get('force_update_or', fuerza)
    
    with st.spinner("📊 Procesando datos Orientadores..."):
        df, resumen = procesar_datos_RP_orientadores(fuerza_actualizacion=fuerza_or)
        
        if 'force_update_or' in st.session_state:
            st.session_state.force_update_or = False
    
    if df is not None:
        if resumen is not None:
            mostrar_tabla_RP_orientadores(resumen)
        else:
            st.error("❌ No se pudieron procesar los datos Orientadores")
    else:
        st.error("❌ No se pudo cargar el archivo de datos")
    
    st.divider()
    
    # =========================================================================
    # 4. PRIMARIA BÁSICA MEDIA
    # =========================================================================
    st.subheader("👩🏻‍🏫 👨🏻‍🏫 Primaria Básica Media")
    
    # Botón específico para esta sección
    col_pbm, _ = st.columns([1, 3])
    with col_pbm:
        if st.button("🔄 Actualizar PBM", key="actualizar_pbm", use_container_width=True):
            st.cache_data.clear()
            st.session_state.force_update_pbm = True
            st.rerun()
    
    fuerza_pbm = st.session_state.get('force_update_pbm', fuerza)
    
    with st.spinner("📊 Procesando datos Primaria Básica Media..."):
        df, resumen = procesar_datos_RP_primaria_basica_media(fuerza_actualizacion=fuerza_pbm)
        
        if 'force_update_pbm' in st.session_state:
            st.session_state.force_update_pbm = False
    
    if df is not None:
        if resumen is not None:
            mostrar_tabla_RP_PBM(resumen)
        else:
            st.error("❌ No se pudieron procesar los datos PBM")
    else:
        st.error("❌ No se pudo cargar el archivo de datos")
    
        # =========================================================================
    # BOTONES DE DESCARGA EXCEL CON FORMATOS
    # =========================================================================
    st.divider()
    st.subheader("📥 Exportar a Excel (con formatos)")
    
    # Crear columnas para botones de descarga
    col_descarga1, col_descarga2, col_descarga3, col_descarga4 = st.columns(4)
    
    # Obtener los dataframes procesados
    with st.spinner("Preparando datos con formatos..."):
        # Recargar o usar datos en cache
        df, resumen_total = procesar_datos_RP_principal()
        _, resumen_pi = procesar_datos_RP_primera_infancia()
        _, resumen_or = procesar_datos_RP_orientadores()
        _, resumen_pbm = procesar_datos_RP_primaria_basica_media()
    
    with col_descarga1:
        if resumen_total is not None:
            excel_total = exportar_a_excel_formateado([resumen_total], ["TOTAL_DOCENTES"], ["RP"])
            st.download_button(
                label="📥 Descargar TOTAL",
                data=excel_total,
                file_name=f"RP_TOTAL_DOCENTES_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                help="Incluye colores y formatos como en pantalla"
            )
        else:
            st.button("📥 Descargar TOTAL", disabled=True, use_container_width=True)
    
    with col_descarga2:
        if resumen_pi is not None:
            excel_pi = exportar_a_excel_formateado([resumen_pi], ["PRIMERA_INFANCIA"], ["RP"])
            st.download_button(
                label="📥 Descargar Primera Infancia",
                data=excel_pi,
                file_name=f"RP_PRIMERA_INFANCIA_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                help="Incluye colores y formatos como en pantalla"
            )
        else:
            st.button("📥 Descargar Primera Infancia", disabled=True, use_container_width=True)
    
    with col_descarga3:
        if resumen_or is not None:
            excel_or = exportar_a_excel_formateado([resumen_or], ["ORIENTADORES"], ["RP"])
            st.download_button(
                label="📥 Descargar Orientadores",
                data=excel_or,
                file_name=f"RP_ORIENTADORES_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                help="Incluye colores y formatos como en pantalla"
            )
        else:
            st.button("📥 Descargar Orientadores", disabled=True, use_container_width=True)
    
    with col_descarga4:
        if resumen_pbm is not None:
            excel_pbm = exportar_a_excel_formateado([resumen_pbm], ["PRIMARIA_BASICA_MEDIA"], ["RP"])
            st.download_button(
                label="📥 Descargar PBM",
                data=excel_pbm,
                file_name=f"RP_PRIMARIA_BASICA_MEDIA_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                help="Incluye colores y formatos como en pantalla"
            )
        else:
            st.button("📥 Descargar PBM", disabled=True, use_container_width=True)
    
    # Botón para descargar TODO en un solo archivo
    st.markdown("---")
    col_todo, _ = st.columns([1, 3])
    with col_todo:
        if all(r is not None for r in [resumen_total, resumen_pi, resumen_or, resumen_pbm]):
            excel_completo = exportar_a_excel_formateado(
                [resumen_total, resumen_pi, resumen_or, resumen_pbm],
                ["TOTAL", "PRIMERA_INFANCIA", "ORIENTADORES", "PRIMARIA_BASICA_MEDIA"],
                ["RP", "RP", "RP", "RP"]
            )
            st.download_button(
                label="📦 Descargar TODO (4 hojas)",
                data=excel_completo,
                file_name=f"RP_DOCENTES_COMPLETO_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
                help="Incluye todas las tablas con sus formatos originales"
            )
        else:
            st.button("📦 Descargar TODO (4 hojas)", disabled=True, use_container_width=True)
    
    # =========================================================================
    # INFO DE ACTUALIZACIÓN FINAL
    # =========================================================================
    st.divider()
    col_info1, col_info2 = st.columns(2)
    with col_info1:
        st.caption(f"📅 {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    with col_info2:
        if 'force_update_all' in st.session_state:
            st.session_state.force_update_all = False
        st.caption("✅ Datos actualizados correctamente" if fuerza else "📊 Usando datos en caché")

# =============================================================================
# PANTALLA DOCENTES SGP - CON 3 TABLAS DETALLADAS
# =============================================================================

def mostrar_pantalla_sgp():
    if st.button("← Volver al Inicio", key="volver_sgp"):
        st.session_state.pagina_actual = "INICIO"
        st.rerun()

    st.markdown("<div class='contenedor-logos'>", unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 2, 1])

    with col1:
        st.image("logo_bogota.png", width=150, use_container_width=True)

    with col2:
        st.markdown(
            "<div class='header'><h2>SISTEMA GENERAL DE PARTICIPACIONES</h2></div>",
            unsafe_allow_html=True
        )

    with col3:
        st.image("logo_alcaldía_mayor.png", width=150, use_container_width=True)

    st.markdown("</div>", unsafe_allow_html=True)

    # Mostrar fecha de actualización
    mostrar_fecha_actualizacion()

    # =========================================================================
    # BOTÓN DE ACTUALIZACIÓN GLOBAL
    # =========================================================================
    col_actualizar, _ = st.columns([1, 3])
    with col_actualizar:
        if st.button("🔄 Actualizar TODOS los Datos SGP", key="actualizar_todos_sgp", use_container_width=True):
            st.cache_data.clear()
            st.session_state.force_update_sgp = True
            st.rerun()

    # Usar la bandera para todas las cargas
    fuerza = st.session_state.get('force_update_sgp', False)

    # =========================================================================
    # 1. TABLA TOTAL (resumen general)
    # =========================================================================
    st.subheader("🌐 TABLA RESUMEN EJECUCIÓN PRESUPUESTAL DOCENTES SGP")

    col_total, _ = st.columns([1, 3])
    with col_total:
        if st.button("🔄 Actualizar TOTAL", key="actualizar_total_sgp", use_container_width=True):
            st.cache_data.clear()
            st.session_state.force_update_total_sgp = True
            st.rerun()

    fuerza_total = st.session_state.get('force_update_total_sgp', fuerza)

    with st.spinner("📊 Procesando datos TOTAL SGP..."):
        df, resumen = procesar_datos_SGP_principal(fuerza_actualizacion=fuerza_total)
        
        if 'force_update_total_sgp' in st.session_state:
            st.session_state.force_update_total_sgp = False

    if df is not None:
        if resumen is not None:
            mostrar_tabla_SGPP_principal(resumen)
        else:
            st.error("❌ No se pudieron procesar los datos TOTAL SGP")
    else:
        st.error("❌ No se pudo cargar el archivo de datos")

    st.divider()

    # =========================================================================
    # 2. TABLA PRIMERA INFANCIA (detallada)
    # =========================================================================
    st.subheader("👶🏻 EJECUCIÓN PRESUPUESTAL DOCENTES PRIMERA INFANCIA SGP")

    col_pi, _ = st.columns([1, 3])
    with col_pi:
        if st.button("🔄 Actualizar Primera Infancia SGP", key="actualizar_pi_sgp", use_container_width=True):
            st.cache_data.clear()
            st.session_state.force_update_pi_sgp = True
            st.rerun()

    fuerza_pi = st.session_state.get('force_update_pi_sgp', fuerza)

    with st.spinner("📊 Procesando datos Primera Infancia SGP..."):
        df, resumen_pi = procesar_datos_SGP_primera_infancia_detallada(fuerza_actualizacion=fuerza_pi)
        
        if 'force_update_pi_sgp' in st.session_state:
            st.session_state.force_update_pi_sgp = False

    if df is not None:
        if resumen_pi is not None:
            mostrar_tabla_SGP_primera_infancia_detallada(resumen_pi)
        else:
            st.error("❌ No se pudieron procesar los datos Primera Infancia SGP")
    else:
        st.error("❌ No se pudo cargar el archivo de datos")

    st.divider()

    # =========================================================================
    # 3. TABLA PRIMARIA BÁSICA Y MEDIA (detallada)
    # =========================================================================
    st.subheader("👩🏻‍🏫 👨🏻‍🏫 EJECUCIÓN PRESUPUESTAL DOCENTES PRIMARIA BÁSICA MEDIA SGP")

    col_pbm, _ = st.columns([1, 3])
    with col_pbm:
        if st.button("🔄 Actualizar PBM SGP", key="actualizar_pbm_sgp", use_container_width=True):
            st.cache_data.clear()
            st.session_state.force_update_pbm_sgp = True
            st.rerun()

    fuerza_pbm = st.session_state.get('force_update_pbm_sgp', fuerza)

    with st.spinner("📊 Procesando datos Primaria Básica y Media SGP..."):
        df, resumen_pbm = procesar_datos_SGP_primaria_basica_media_detallada(fuerza_actualizacion=fuerza_pbm)
        
        if 'force_update_pbm_sgp' in st.session_state:
            st.session_state.force_update_pbm_sgp = False

    if df is not None:
        if resumen_pbm is not None:
            mostrar_tabla_SGP_primaria_basica_media_detallada(resumen_pbm)
        else:
            st.error("❌ No se pudieron procesar los datos PBM SGP")
    else:
        st.error("❌ No se pudo cargar el archivo de datos")

    # =========================================================================
    # BOTONES DE DESCARGA EXCEL
    # =========================================================================
    st.divider()
    st.subheader("📥 Exportar a Excel")

    # Crear columnas para botones de descarga
    col_descarga1, col_descarga2, col_descarga3, col_descarga4 = st.columns(4)

    # Obtener los dataframes procesados
    with st.spinner("Preparando datos con formatos..."):
        # Recargar o usar datos en cache
        _, resumen_total = procesar_datos_SGP_principal()
        _, resumen_pi = procesar_datos_SGP_primera_infancia_detallada()
        _, resumen_pbm = procesar_datos_SGP_primaria_basica_media_detallada()

    with col_descarga1:
        if resumen_total is not None:
            excel_total = exportar_a_excel_formateado([resumen_total], ["TOTAL_SGP"], ["SGP"])
            st.download_button(
                label="📥 Descargar TOTAL",
                data=excel_total,
                file_name=f"SGP_TOTAL_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                help="Incluye colores y formatos como en pantalla"
            )
        else:
            st.button("📥 Descargar TOTAL", disabled=True, use_container_width=True)

    with col_descarga2:
        if resumen_pi is not None:
            excel_pi = exportar_a_excel_formateado([resumen_pi], ["PRIMERA_INFANCIA_SGP"], ["SGP"])
            st.download_button(
                label="📥 Descargar Primera Infancia",
                data=excel_pi,
                file_name=f"SGP_PRIMERA_INFANCIA_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                help="Incluye colores y formatos como en pantalla"
            )
        else:
            st.button("📥 Descargar Primera Infancia", disabled=True, use_container_width=True)

    with col_descarga3:
        if resumen_pbm is not None:
            excel_pbm = exportar_a_excel_formateado([resumen_pbm], ["PRIMARIA_BASICA_MEDIA_SGP"], ["SGP"])
            st.download_button(
                label="📥 Descargar PBM",
                data=excel_pbm,
                file_name=f"SGP_PBM_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                help="Incluye colores y formatos como en pantalla"
            )
        else:
            st.button("📥 Descargar PBM", disabled=True, use_container_width=True)

    with col_descarga4:
        # Descargar TODO en un solo archivo
        if all(r is not None for r in [resumen_total, resumen_pi, resumen_pbm]):
            excel_completo = exportar_a_excel_formateado(
                [resumen_total, resumen_pi, resumen_pbm],
                ["TOTAL_SGP", "PRIMERA_INFANCIA_SGP", "PRIMARIA_BASICA_MEDIA_SGP"],
                ["SGP", "SGP", "SGP"]
            )
            st.download_button(
                label="📦 Descargar TODO (3 hojas)",
                data=excel_completo,
                file_name=f"SGP_COMPLETO_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
                help="Incluye todas las tablas SGP con sus formatos originales"
            )
        else:
            st.button("📦 Descargar TODO (3 hojas)", disabled=True, use_container_width=True)

    # =========================================================================
    # INFO DE ACTUALIZACIÓN FINAL
    # =========================================================================
    st.divider()
    col_info1, col_info2 = st.columns(2)
    with col_info1:
        fecha = st.session_state.get('fecha_actualizacion', datetime.now().strftime('%d/%m/%Y %H:%M'))
        st.caption(f"📅 {fecha}")
    with col_info2:
        if 'force_update_sgp' in st.session_state:
            st.session_state.force_update_sgp = False
        st.caption("✅ Datos actualizados correctamente" if fuerza else "📊 Usando datos en caché")

# =============================================================================
# MAIN - MODIFICADO PARA PRECARGAR DATOS
# =============================================================================
def main():
    # Inicializar session state
    if "pagina_actual" not in st.session_state:
        st.session_state.pagina_actual = "INICIO"
    
    # 🔹 NUEVO: Precargar datos al inicio para tener la fecha disponible
    if "fecha_actualizacion" not in st.session_state or st.session_state.fecha_actualizacion == "No disponible":
        with st.spinner("🔄 Cargando datos iniciales..."):
            try:
                # Hacer una carga ligera solo para obtener la fecha
                df = cargar_datos_originales(_fuerza_actualizacion=False)
                if df is not None:
                    # La fecha ya se guardó en session_state dentro de cargar_datos_originales
                    pass
            except Exception as e:
                # Si falla, establecer un valor por defecto
                st.session_state.fecha_actualizacion = "Cargando..."
    
    # Cargar estilos
    cargar_estilos()
    
    # Navegación
    if st.session_state.pagina_actual == "INICIO":
        mostrar_pantalla_inicial()
    elif st.session_state.pagina_actual == "POR_FUENTE":
        mostrar_pantalla_por_fuente()
    elif st.session_state.pagina_actual == "RECURSOS_PROPIOS":
        mostrar_pantalla_recursos_propios()
    elif st.session_state.pagina_actual == "SGP":
        mostrar_pantalla_sgp()

if __name__ == "__main__":
    main()

